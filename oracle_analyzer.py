#!/usr/bin/env python3
"""
Script per analisi dipendenze Oracle e stima migrazione PostgreSQL
Versione con prefissi tabelle pdt_dep_dba_/pdt_dep_nodba_, pdt_sizes_dba_/pdt_sizes_nodba_, rilevamento privilegi DBA,
configurazioni esterne multiple, gestione connessioni normalizzata, output Excel e analisi dimensioni
ORA2PG: Utilizza schema specificato nel config JSON, altrimenti logica DBA/NON-DBA
Query tablespace aggiornate: DBA (dettagli completi), NON-DBA (aggregato per tablespace)
MODIFICHE: File Excel specifici disattivati, naming con schema invece di username
"""

import oracledb
import psycopg2
import os
import csv
import json
import subprocess
import re
from datetime import datetime
from pathlib import Path
import sys
import logging
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Fix encoding per Windows
import locale
if sys.platform == 'win32':
    # Forza UTF-8 su Windows
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

class OracleMultiDatabaseAnalyzer:
    def __init__(self, config_file="oracle_connections.json"):
        print(f"🚀 Inizializzazione OracleMultiDatabaseAnalyzer...")
        print(f"📄 File configurazione: {config_file}")
        
        self.config_file = config_file
        self.config = self.load_configuration()
        self.pg_config = self.config['postgresql_config']
        self.oracle_connections = self.config['oracle_connections']
        
        # Configurazione Oracle Client
        self.setup_oracle_client()
        
        # Directory output
        self.output_dir = f"oracle_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        os.makedirs(self.output_dir, exist_ok=True)
        print(f"📁 Directory output: {self.output_dir}")
        
        # Configurazione output - DISABILITA CSV, ABILITA EXCEL
        self.generate_csv = False      # 🔴 CSV DISABILITATO
        self.generate_excel = True     # 🟢 EXCEL ABILITATO
        
        # 🆕 CONFIGURAZIONE ANALISI DIMENSIONI
        self.analyze_sizes = self.config.get('analyze_sizes', True)  # Default abilitato
        
        # 🆕 CONFIGURAZIONE OUTPUT ORA2PG
        # Opzioni: 'html_only', 'html_and_txt'
        self.ora2pg_output_mode = self.config.get('ora2pg_output_mode', 'html_and_txt')
        
        # Log delle connessioni caricate
        for conn in self.oracle_connections:
            schema_info = f" -> schema: {conn['schema']}" if 'schema' in conn else " -> schema: auto (DBA/NON-DBA)"
            print(f"  - {conn['connection_name']}: {conn['user']}@{conn.get('dsn', 'N/A')}{schema_info}")
    
    def setup_oracle_client(self):
        """Configura Oracle Client con path automatico o configurabile"""
        try:
            # Prova a caricare Oracle Client dai path comuni
            oracle_client_paths = []
            
            if sys.platform == 'win32':
                oracle_client_paths = [
                    r"C:\instantclient_23_7\instantclient_23_7",
                    r"C:\instantclient_21_7",
                    r"C:\instantclient_19_8",
                    r"C:\oracle\instantclient_23_7",
                    r"C:\oracle\instantclient_21_7"
                ]
            else:
                oracle_client_paths = [
                    "/usr/lib/oracle/21/client64/lib",
                    "/usr/lib/oracle/19.8/client64/lib",
                    "/opt/oracle/instantclient_21_7",
                    "/opt/oracle/instantclient_19_8"
                ]
            
            # Prova path dalla configurazione
            if 'oracle_client_path' in self.config:
                oracle_client_paths.insert(0, self.config['oracle_client_path'])
            
            # Prova a inizializzare Oracle Client
            oracle_initialized = False
            
            for path in oracle_client_paths:
                if os.path.exists(path):
                    try:
                        oracledb.init_oracle_client(lib_dir=path)
                        print(f"✅ Oracle Client inizializzato da: {path}")
                        oracle_initialized = True
                        break
                    except Exception as e:
                        print(f"⚠️  Tentativo fallito per {path}: {e}")
                        continue
            
            # Prova senza specificare path (se Oracle è nel PATH di sistema)
            if not oracle_initialized:
                try:
                    oracledb.init_oracle_client()
                    print("✅ Oracle Client inizializzato dal PATH di sistema")
                    oracle_initialized = True
                except Exception as e:
                    print(f"❌ Impossibile inizializzare Oracle Client: {e}")
                    print("💡 Suggerimenti:")
                    print("   1. Installa Oracle Instant Client")
                    print("   2. Aggiungi 'oracle_client_path' alla configurazione JSON")
                    print("   3. Aggiungi Oracle Client al PATH di sistema")
                    raise
            
        except Exception as e:
            print(f"❌ ERRORE CRITICO Oracle Client: {e}")
            raise
    
    def load_configuration(self):
        """Carica la configurazione dal file JSON"""
        try:
            print(f"📖 Caricamento configurazione da: {self.config_file}")
            
            with open(self.config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # Validazione configurazione
            if 'postgresql_config' not in config:
                raise ValueError("Configurazione PostgreSQL mancante")
            if 'oracle_connections' not in config:
                raise ValueError("Configurazioni Oracle mancanti")
            if not config['oracle_connections']:
                raise ValueError("Nessuna connessione Oracle definita")
            
            # Validazione campi obbligatori per ogni connessione
            required_fields = ['connection_name', 'user', 'password']
            for i, conn in enumerate(config['oracle_connections']):
                for field in required_fields:
                    if field not in conn:
                        raise ValueError(f"Campo '{field}' mancante nella connessione {i+1}")
                
                # DSN non è obbligatorio se non specificato
                if 'dsn' not in conn:
                    print(f"⚠️  Connessione {conn['connection_name']}: DSN mancante, verrà saltata")
            
            print("✅ Configurazione validata con successo")
            return config
            
        except FileNotFoundError:
            print(f"❌ ERRORE: File di configurazione '{self.config_file}' non trovato!")
            print("🔧 Creo un file di esempio...")
            self.create_sample_config()
            sys.exit(1)
        except json.JSONDecodeError as e:
            print(f"❌ ERRORE: Formato JSON non valido nel file '{self.config_file}': {e}")
            sys.exit(1)
        except Exception as e:
            print(f"❌ ERRORE: {e}")
            sys.exit(1)
    
    def create_sample_config(self):
        """Crea un file di configurazione di esempio"""
        sample_config = {
            "postgresql_config": {
                "host": "localhost",
                "database": "DBGRMED",
                "user": "postgres",
                "password": "your_pg_password",
                "port": 5432
            },
            "oracle_connections": [
                {
                    "connection_name": "GRMED_PROD",
                    "dsn": "10.138.154.6:10461/GRMED",
                    "user": "GRMED",
                    "password": "your_oracle_password",
                    "schema": "GRMED",
                    "description": "Database GRMED produzione",
                    "is_dba": "auto",
                    "analyze_all_schemas": True
                },
                {
                    "connection_name": "EXAMPLE_DB",
                    "dsn": "hostname:port/service_name",
                    "user": "username",
                    "password": "password",
                    "schema": "TARGET_SCHEMA",
                    "description": "Descrizione database di esempio",
                    "is_dba": False,
                    "analyze_all_schemas": False
                }
            ],
            "oracle_client_path": "C:/instantclient_23_7/instantclient_23_7",
            "ora2pg_output_mode": "html_and_txt",
            "analyze_sizes": True
        }

        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(sample_config, f, indent=2, ensure_ascii=False)
        
        print(f"📄 File di esempio creato: {self.config_file}")
        print("✏️  Modifica il file con le tue configurazioni e riesegui lo script.")
        print("💡 Opzioni 'is_dba': true, false, 'auto' (rileva automaticamente)")
        print("💡 Opzioni 'analyze_all_schemas': true (tutti gli schemi), false (solo utente corrente)")
        print("💡 Campo 'schema': opzionale, specifica lo schema per ora2pg (indipendente da DBA/NON-DBA)")
        
    def get_db_connection(self, dsn, user, password):
        """Connessione al database Oracle usando oracledb"""
        try:
            print(f"🔗 Tentativo connessione Oracle: {user}@{dsn}")
            connection = oracledb.connect(user=user, password=password, dsn=dsn)
            print(f"✅ Connessione Oracle riuscita")
            return connection
        except Exception as e:
            print(f"❌ ERRORE connessione Oracle {user}@{dsn}: {str(e)}")
            raise
    
    def check_dba_privileges(self, connection, db_config):
        """🆕 Verifica se l'utente ha privilegi DBA"""
        cursor = connection.cursor()
        
        # Ottieni l'utente corrente
        cursor.execute("SELECT USER FROM DUAL")
        current_user = cursor.fetchone()[0]
        
        # Controlla configurazione manuale
        if 'is_dba' in db_config:
            if db_config['is_dba'] == 'auto':
                # Rilevamento automatico
                pass  # Continua con i test sotto
            elif isinstance(db_config['is_dba'], bool):
                # Configurazione manuale
                is_dba = db_config['is_dba']
                print(f"    🔧 Privilegi DBA configurati manualmente: {'SÌ' if is_dba else 'NO'}")
                cursor.close()
                return is_dba
        
        # Rilevamento automatico privilegi DBA
        print(f"    🔍 Rilevamento automatico privilegi DBA per {current_user}...")
        
        dba_tests = [
            # Test 1: Verifica se utente ha ruolo DBA
            ("SELECT COUNT(*) FROM session_roles WHERE role = 'DBA'", "Ruolo DBA"),
            
            # Test 2: Verifica accesso a DBA_USERS
            ("SELECT COUNT(*) FROM dba_users WHERE rownum = 1", "Accesso DBA_USERS"),
            
            # Test 3: Verifica se utente è SYS o SYSTEM
            (f"SELECT CASE WHEN USER IN ('SYS', 'SYSTEM') THEN 1 ELSE 0 END FROM DUAL", "Utente amministrativo"),
            
            # Test 4: Verifica privilegi SELECT ANY TABLE
            ("SELECT COUNT(*) FROM session_privs WHERE privilege = 'SELECT ANY TABLE'", "Privilegio SELECT ANY TABLE")
        ]
        
        dba_score = 0
        total_tests = len(dba_tests)
        
        for test_query, test_name in dba_tests:
            try:
                cursor.execute(test_query)
                result = cursor.fetchone()[0]
                if result > 0:
                    print(f"      ✅ {test_name}: PASS")
                    dba_score += 1
                else:
                    print(f"      ❌ {test_name}: FAIL")
            except Exception as e:
                print(f"      ⚠️  {test_name}: ERRORE ({str(e)})")
        
        # Se l'utente passa almeno 2 test su 4, considerarlo DBA
        is_dba = dba_score >= 2
        
        print(f"    📊 Score DBA: {dba_score}/{total_tests}")
        print(f"    🎯 Privilegi DBA rilevati: {'SÌ' if is_dba else 'NO'}")
        
        cursor.close()
        return is_dba
    
    def get_oracle_dependencies_dba(self, connection, db_config):
        """🆕 Estrae dipendenze per utenti DBA (tutti gli schemi del database)"""
        cursor = connection.cursor()
        
        # Ottieni l'utente corrente
        cursor.execute("SELECT USER FROM DUAL")
        current_user = cursor.fetchone()[0]
        print(f"    📊 Analisi dipendenze DBA per utente: {current_user}")
        
        # Verifica se analizzare tutti gli schemi
        analyze_all = db_config.get('analyze_all_schemas', True)
        schema_filter = ""
        params = {}
        
        if not analyze_all:
            # Solo schema dell'utente corrente
            schema_filter = "AND (d.owner = :current_user OR d.referenced_owner = :current_user)"
            params['current_user'] = current_user
            print(f"    🎯 Modalità: Solo schema {current_user}")
        else:
            print(f"    🎯 Modalità: Tutti gli schemi del database")
        
        # Query per dipendenze tra schemi (DBA - accesso completo)
        dependencies_query = f"""
        SELECT 
            d.owner AS source_owner,
            d.name AS source_name,
            d.type AS source_type,
            d.referenced_owner AS target_owner,
            d.referenced_name AS target_name,
            d.referenced_type AS target_type,
            d.referenced_link_name AS db_link
        FROM 
            dba_dependencies d
        WHERE 
            d.owner <> d.referenced_owner
            AND d.referenced_owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')
            AND d.owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')
            {schema_filter}
        ORDER BY 
            d.owner, d.name
        """
        
        # Query per DB Links accessibili (DBA - tutti i DB links)
        dblinks_query = f"""
        SELECT 
            owner,
            db_link,
            username,
            host
        FROM 
            dba_db_links
        WHERE 
            owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC')
            {("AND owner = :current_user" if not analyze_all else "")}
        ORDER BY owner, db_link
        """
        
        # Query per oggetti degli schemi (DBA)
        objects_query = f"""
        SELECT 
            owner,
            object_type,
            COUNT(*) as object_count
        FROM 
            dba_objects
        WHERE 
            object_type NOT LIKE '%PARTITION%'
            AND owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')
            {("AND owner = :current_user" if not analyze_all else "")}
        GROUP BY 
            owner, object_type
        ORDER BY 
            owner, object_type
        """
        
        # 🔧 QUERY CORRETTA: Usa 'owner' invece di 'table_schema' per dba_tab_privs
        cross_schema_query = f"""
        SELECT DISTINCT
            p.grantor AS privilege_grantor,
            p.grantee AS privilege_grantee,
            p.owner AS table_schema,
            p.table_name,
            p.privilege
        FROM 
            dba_tab_privs p
        WHERE 
            p.grantee NOT IN ('PUBLIC', 'SYS', 'SYSTEM', 'OUTLN', 'DBSNMP')
            AND p.grantor NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')
            AND p.grantor <> p.grantee
            {("AND (p.grantor = :current_user OR p.grantee = :current_user)" if not analyze_all else "")}
        ORDER BY p.grantee, p.owner, p.table_name, p.privilege
        """
        
        # Query per oggetti referenziati da altri schemi (DBA)
        external_refs_query = f"""
        SELECT DISTINCT
            s.owner AS synonym_owner,
            s.synonym_name,
            s.table_owner AS referenced_owner,
            s.table_name AS referenced_object,
            s.db_link
        FROM 
            dba_synonyms s
        WHERE 
            s.owner != s.table_owner
            AND s.owner NOT IN ('PUBLIC', 'SYS', 'SYSTEM', 'OUTLN', 'DBSNMP')
            AND s.table_owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')
            {("AND s.table_owner = :current_user" if not analyze_all else "")}
        ORDER BY s.owner, s.synonym_name, s.table_owner, s.table_name, s.db_link
        """
        
        return self._execute_dependency_queries(cursor, {
            'dependencies': dependencies_query,
            'db_links': dblinks_query,
            'object_summary': objects_query,
            'cross_schema_privs': cross_schema_query,
            'external_references': external_refs_query
        }, params, is_dba=True)
    
    def get_oracle_dependencies_non_dba(self, connection):
        """🆕 Estrae dipendenze per utenti NON DBA (solo schema corrente)"""
        cursor = connection.cursor()
        
        # Ottieni l'utente corrente
        cursor.execute("SELECT USER FROM DUAL")
        current_user = cursor.fetchone()[0]
        print(f"    📊 Analisi dipendenze NON-DBA per utente: {current_user}")
        
        # Query per dipendenze tra schemi (NON DBA - solo ALL_DEPENDENCIES)
        dependencies_query = """
        SELECT 
            d.owner AS source_owner,
            d.name AS source_name,
            d.type AS source_type,
            d.referenced_owner AS target_owner,
            d.referenced_name AS target_name,
            d.referenced_type AS target_type,
            d.referenced_link_name AS db_link
        FROM 
            all_dependencies d
        WHERE 
            (d.owner = :current_user OR d.referenced_owner = :current_user)
            AND d.owner <> d.referenced_owner
            AND d.referenced_owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC')
        ORDER BY 
            d.owner, d.name
        """
        
        # Query per DB Links accessibili all'utente (NON DBA)
        dblinks_query = """
        SELECT 
            owner,
            db_link,
            username,
            host
        FROM 
            all_db_links
        WHERE 
            owner = :current_user
            OR owner = 'PUBLIC'
        """
        
        # Query per oggetti dello schema corrente (NON DBA)
        objects_query = """
        SELECT 
            owner,
            object_type,
            COUNT(*) as object_count
        FROM 
            all_objects
        WHERE 
            owner = :current_user
            AND object_type NOT LIKE '%PARTITION%'
        GROUP BY 
            owner, object_type
        ORDER BY 
            owner, object_type
        """
        
        # 🔧 QUERY CORRETTA: Usa 'owner' invece di 'table_schema' per all_tab_privs
        cross_schema_query = """
        SELECT DISTINCT
            p.grantor AS privilege_grantor,
            p.grantee AS privilege_grantee,
            p.owner AS table_schema,
            p.table_name,
            p.privilege
        FROM 
            all_tab_privs p
        WHERE 
            (p.grantor = :current_user OR p.grantee = :current_user)
            AND p.grantee NOT IN ('PUBLIC', 'SYS', 'SYSTEM')
        ORDER BY p.grantee, p.owner, p.table_name, p.privilege
        """
        
        # Query per oggetti referenziati da altri schemi (NON DBA)
        external_refs_query = """
        SELECT DISTINCT
            s.owner AS synonym_owner,
            s.synonym_name,
            s.table_owner AS referenced_owner,
            s.table_name AS referenced_object,
            s.db_link
        FROM 
            all_synonyms s
        WHERE 
            s.table_owner = :current_user
            AND s.owner != :current_user
            AND s.owner NOT IN ('PUBLIC', 'SYS', 'SYSTEM')
        ORDER BY s.owner, s.synonym_name, s.table_owner, s.table_name, s.db_link
        """
        
        return self._execute_dependency_queries(cursor, {
            'dependencies': dependencies_query,
            'db_links': dblinks_query,
            'object_summary': objects_query,
            'cross_schema_privs': cross_schema_query,
            'external_references': external_refs_query
        }, {'current_user': current_user}, is_dba=False)
    
    def _execute_dependency_queries(self, cursor, queries, params, is_dba):
        """🆕 Esegue le query delle dipendenze e gestisce i risultati"""
        results = {
            'dependencies': [],
            'db_links': [],
            'object_summary': [],
            'cross_schema_privs': [],
            'external_references': [],
            'is_dba': is_dba
        }
        
        query_descriptions = {
            'dependencies': 'Dipendenze',
            'db_links': 'DB Links',
            'object_summary': 'Oggetti',
            'cross_schema_privs': 'Privilegi cross-schema',
            'external_references': 'Riferimenti esterni'
        }
        
        for query_type, query_sql in queries.items():
            try:
                if params:
                    cursor.execute(query_sql, **params)
                else:
                    cursor.execute(query_sql)
                
                results[query_type] = cursor.fetchall()
                count = len(results[query_type])
                print(f"    - {query_descriptions[query_type]}: {count} record")
                
            except Exception as e:
                print(f"    ⚠️  Errore query {query_descriptions[query_type]}: {str(e)}")
                results[query_type] = []
        
        cursor.close()
        return results
    
    def get_oracle_dependencies(self, connection, is_dba, db_config):
        """🆕 Dispatcher per query dipendenze in base ai privilegi"""
        if is_dba:
            return self.get_oracle_dependencies_dba(connection, db_config)
        else:
            return self.get_oracle_dependencies_non_dba(connection)
    
    def get_oracle_sizes_dba(self, connection, db_config):
        """🆕 Estrae informazioni dimensioni per utenti DBA"""
        cursor = connection.cursor()
        
        # Ottieni l'utente corrente
        cursor.execute("SELECT USER FROM DUAL")
        current_user = cursor.fetchone()[0]
        print(f"    📏 Analisi dimensioni DBA per utente: {current_user}")
        
        # Verifica se analizzare tutti gli schemi
        analyze_all = db_config.get('analyze_all_schemas', True)
        schema_filter = ""
        params = {}
        
        if not analyze_all:
            schema_filter = "WHERE owner = :current_user"
            params['current_user'] = current_user
            print(f"    🎯 Dimensioni modalità: Solo schema {current_user}")
        else:
            schema_filter = "WHERE owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')"
            print(f"    🎯 Dimensioni modalità: Tutti gli schemi del database")
        
        # Query per dimensioni database (DBA)
        database_size_query = """
        SELECT 
            'DATABASE_SIZE' as metric_type,
            'TOTAL' as object_name,
            ROUND(SUM(bytes)/1024/1024/1024, 2) as size_gb,
            ROUND(SUM(bytes)/1024/1024, 2) as size_mb,
            SUM(bytes) as size_bytes,
            COUNT(*) as file_count
        FROM 
            dba_data_files
        UNION ALL
        SELECT 
            'TEMP_SIZE' as metric_type,
            'TEMP' as object_name,
            ROUND(SUM(bytes)/1024/1024/1024, 2) as size_gb,
            ROUND(SUM(bytes)/1024/1024, 2) as size_mb,
            SUM(bytes) as size_bytes,
            COUNT(*) as file_count
        FROM 
            dba_temp_files
        """
        
        # 🆕 Query per dimensioni tablespace (DBA) - NUOVA QUERY DETTAGLIATA
        tablespace_size_query = """
        SELECT
            ts.tablespace_name,
            ts.status,
            ts.contents as type,
            -- Spazio allocato (dimensione fisica dei file)
            ROUND(NVL(df.allocated_gb, 0), 2) as allocated_gb,
            ROUND(NVL(df.allocated_mb, 0), 2) as allocated_mb,
            df.allocated_bytes,
            -- Spazio utilizzato (dai segmenti)
            ROUND(NVL(seg.used_gb, 0), 2) as used_gb,
            ROUND(NVL(seg.used_mb, 0), 2) as used_mb,
            seg.used_bytes,
            -- Spazio libero
            ROUND(NVL(fs.free_gb, 0), 2) as free_gb,
            ROUND(NVL(fs.free_mb, 0), 2) as free_mb,
            fs.free_bytes,
            -- Percentuali
            ROUND(CASE WHEN df.allocated_gb > 0 THEN (seg.used_gb / df.allocated_gb) * 100 ELSE 0 END, 2) as pct_used,
            ROUND(CASE WHEN df.allocated_gb > 0 THEN (fs.free_gb / df.allocated_gb) * 100 ELSE 0 END, 2) as pct_free,
            -- Numero di file e segmenti
            NVL(df.file_count, 0) as datafile_count,
            NVL(seg.segment_count, 0) as segment_count
        FROM
            dba_tablespaces ts
        LEFT JOIN
            -- Spazio allocato dai datafile
            (SELECT
                tablespace_name,
                ROUND(SUM(bytes)/1024/1024/1024, 2) as allocated_gb,
                ROUND(SUM(bytes)/1024/1024, 2) as allocated_mb,
                SUM(bytes) as allocated_bytes,
                COUNT(*) as file_count
             FROM dba_data_files
             GROUP BY tablespace_name) df
        ON ts.tablespace_name = df.tablespace_name
        LEFT JOIN
            -- Spazio libero
            (SELECT
                tablespace_name,
                ROUND(SUM(bytes)/1024/1024/1024, 2) as free_gb,
                ROUND(SUM(bytes)/1024/1024, 2) as free_mb,
                SUM(bytes) AS free_bytes
             FROM dba_free_space
             GROUP BY tablespace_name) fs
        ON ts.tablespace_name = fs.tablespace_name
        LEFT JOIN
            -- Spazio utilizzato dai segmenti
            (SELECT
                tablespace_name,
                ROUND(SUM(bytes)/1024/1024/1024, 2) as used_gb,
                ROUND(SUM(bytes)/1024/1024, 2) as used_mb,
                SUM(bytes) AS used_bytes,
                COUNT(*) as segment_count
             FROM dba_segments
             GROUP BY tablespace_name) seg
        ON ts.tablespace_name = seg.tablespace_name
        WHERE ts.contents != 'TEMPORARY'  -- Escludi tablespace temporanei
        ORDER BY df.tablespace_name NULLS LAST
        """
        
        # Query per dimensioni schema (DBA)
        schema_size_query = f"""
        SELECT 
            owner,
            ROUND(SUM(bytes)/1024/1024/1024, 2) as size_gb,
            ROUND(SUM(bytes)/1024/1024, 2) as size_mb,
            SUM(bytes) as size_bytes,
            COUNT(*) as segment_count
        FROM 
            dba_segments
        {schema_filter}
        GROUP BY 
            owner
        ORDER BY owner
        """
        
        # Query per dimensioni tabelle (DBA)
        table_size_query = f"""
        SELECT 
            owner,
            segment_name as table_name,
            segment_type,
            tablespace_name,
            ROUND(bytes/1024/1024/1024, 4) as size_gb,
            ROUND(bytes/1024/1024, 2) as size_mb,
            bytes as size_bytes,
            blocks,
            extents
        FROM 
            dba_segments
        WHERE 
            segment_type IN ('TABLE', 'TABLE PARTITION', 'TABLE SUBPARTITION')
            {("AND owner = :current_user" if not analyze_all else "AND owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')")}
        ORDER BY 
            owner, segment_name
        """
        
        # Query per dimensioni indici (DBA)
        index_size_query = f"""
        SELECT 
            owner,
            segment_name as index_name,
            segment_type,
            tablespace_name,
            ROUND(bytes/1024/1024/1024, 4) as size_gb,
            ROUND(bytes/1024/1024, 2) as size_mb,
            bytes as size_bytes,
            blocks,
            extents
        FROM 
            dba_segments
        WHERE 
            segment_type LIKE '%INDEX%'
            {("AND owner = :current_user" if not analyze_all else "AND owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')")}
        ORDER BY 
            owner, segment_name
        """
        
        # Query per dimensioni segmenti (DBA)
        segment_size_query = f"""
        SELECT 
            owner,
            segment_name,
            segment_type,
            tablespace_name,
            ROUND(bytes/1024/1024/1024, 4) as size_gb,
            ROUND(bytes/1024/1024, 2) as size_mb,
            bytes as size_bytes,
            blocks,
            extents,
            initial_extent,
            next_extent,
            max_extents
        FROM 
            dba_segments
        {("WHERE owner = :current_user" if not analyze_all else "WHERE owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')")}
        ORDER BY 
            owner, segment_name
        """
        
        # Query per dimensioni codice (DBA - ma sempre limitato agli schemi accessibili)
        code_size_query = f"""
        SELECT 
            owner,
            name as object_name,
            type as object_type,
            LENGTH(text) as char_length,
            LENGTHB(text) as byte_length,
            line,
            text
        FROM 
            all_source
        WHERE 
            type IN ('PACKAGE', 'PACKAGE BODY', 'PROCEDURE', 'FUNCTION', 'TRIGGER')
            {("AND owner = :current_user" if not analyze_all else "AND owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')")}
        ORDER BY 
            owner, name, type, line
        """
        
        # Query aggregata per statistiche codice (DBA)
        code_stats_query = f"""
        SELECT 
            owner,
            name as object_name,
            type as object_type,
            COUNT(*) as total_lines,
            SUM(LENGTH(text)) as total_chars,
            SUM(LENGTHB(text)) as total_bytes,
            MIN(line) as first_line,
            MAX(line) as last_line
        FROM 
            all_source
        WHERE 
            type IN ('PACKAGE', 'PACKAGE BODY', 'PROCEDURE', 'FUNCTION', 'TRIGGER')
            {("AND owner = :current_user" if not analyze_all else "AND owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP')")}
        GROUP BY 
            owner, name, type
        ORDER BY 
            owner, name, type
        """
        
        return self._execute_size_queries(cursor, {
            'database_size': database_size_query,
            'tablespace_size': tablespace_size_query,
            'schema_size': schema_size_query,
            'table_size': table_size_query,
            'index_size': index_size_query,
            'segment_size': segment_size_query,
            'code_lines': code_size_query,
            'code_stats': code_stats_query
        }, params, is_dba=True)
    
    def get_oracle_sizes_non_dba(self, connection):
        """🆕 Estrae informazioni dimensioni per utenti NON DBA"""
        cursor = connection.cursor()
        
        # Ottieni l'utente corrente
        cursor.execute("SELECT USER FROM DUAL")
        current_user = cursor.fetchone()[0]
        print(f"    📏 Analisi dimensioni NON-DBA per utente: {current_user}")
        
        # Query per dimensioni database (NON DBA) - informazioni limitate
        database_size_query = """
        SELECT 
            'USER_OBJECTS' as metric_type,
            'USER_SCHEMA' as object_name,
            ROUND(SUM(bytes)/1024/1024/1024, 2) as size_gb,
            ROUND(SUM(bytes)/1024/1024, 2) as size_mb,
            SUM(bytes) as size_bytes,
            COUNT(*) as segment_count
        FROM 
            user_segments
        """
        
        # 🆕 Query per dimensioni tablespace (NON DBA) - NUOVA QUERY AGGREGATA
        tablespace_size_query = """
        SELECT DISTINCT
            tablespace_name,
            ROUND(SUM(bytes) / 1024 / 1024 / 1024, 2) as used_gb,
            ROUND(SUM(bytes) / 1024 / 1024, 2) as used_mb,
            SUM(bytes) as used_bytes,
            COUNT(*) as file_count,
            'UNKNOWN' as status
        FROM
            user_segments
        GROUP BY
            tablespace_name
        ORDER BY
            tablespace_name NULLS LAST
        """
        
        # Query per dimensioni schema (NON DBA)
        schema_size_query = """
        SELECT 
            USER as owner,
            ROUND(SUM(bytes)/1024/1024/1024, 2) as size_gb,
            ROUND(SUM(bytes)/1024/1024, 2) as size_mb,
            SUM(bytes) as size_bytes,
            COUNT(*) as segment_count
        FROM 
            user_segments
        GROUP BY 
            USER
        """
        
        # Query per dimensioni tabelle (NON DBA)
        table_size_query = """
        SELECT 
            USER as owner,
            segment_name as table_name,
            segment_type,
            tablespace_name,
            ROUND(bytes/1024/1024/1024, 4) as size_gb,
            ROUND(bytes/1024/1024, 2) as size_mb,
            bytes as size_bytes,
            blocks,
            extents
        FROM 
            user_segments
        WHERE 
            segment_type IN ('TABLE', 'TABLE PARTITION', 'TABLE SUBPARTITION')
        ORDER BY 
            owner, segment_name, segment_type
        """
        
        # Query per dimensioni indici (NON DBA)
        index_size_query = """
        SELECT 
            USER as owner,
            segment_name as index_name,
            segment_type,
            tablespace_name,
            ROUND(bytes/1024/1024/1024, 4) as size_gb,
            ROUND(bytes/1024/1024, 2) as size_mb,
            bytes as size_bytes,
            blocks,
            extents
        FROM 
            user_segments
        WHERE 
            segment_type LIKE '%INDEX%'
        ORDER BY 
            owner, segment_name, segment_type
        """
        
        # Query per dimensioni segmenti (NON DBA)
        segment_size_query = """
        SELECT 
            USER as owner,
            segment_name,
            segment_type,
            tablespace_name,
            ROUND(bytes/1024/1024/1024, 4) as size_gb,
            ROUND(bytes/1024/1024, 2) as size_mb,
            bytes as size_bytes,
            blocks,
            extents,
            initial_extent,
            next_extent,
            max_extents
        FROM 
            user_segments
        ORDER BY 
            owner, segment_name, segment_type
        """
        
        # Query per dimensioni codice (NON DBA) - solo USER_SOURCE
        code_size_query = """
        SELECT 
            USER as owner,
            name as object_name,
            type as object_type,
            LENGTH(text) as char_length,
            LENGTHB(text) as byte_length,
            line,
            text
        FROM 
            user_source
        WHERE 
            type IN ('PACKAGE', 'PACKAGE BODY', 'PROCEDURE', 'FUNCTION', 'TRIGGER')
        ORDER BY 
            user, name, type, line
        """
        
        # Query aggregata per statistiche codice (NON DBA)
        code_stats_query = """
        SELECT 
            USER as owner,
            name as object_name,
            type as object_type,
            COUNT(*) as total_lines,
            SUM(LENGTH(text)) as total_chars,
            SUM(LENGTHB(text)) as total_bytes,
            MIN(line) as first_line,
            MAX(line) as last_line
        FROM 
            user_source
        WHERE 
            type IN ('PACKAGE', 'PACKAGE BODY', 'PROCEDURE', 'FUNCTION', 'TRIGGER')
        GROUP BY 
            name, type
        ORDER BY 
            user, name, type
        """
        
        return self._execute_size_queries(cursor, {
            'database_size': database_size_query,
            'tablespace_size': tablespace_size_query,
            'schema_size': schema_size_query,
            'table_size': table_size_query,
            'index_size': index_size_query,
            'segment_size': segment_size_query,
            'code_lines': code_size_query,
            'code_stats': code_stats_query
        }, {}, is_dba=False)
    
    def _execute_size_queries(self, cursor, queries, params, is_dba):
        """🆕 Esegue le query delle dimensioni e gestisce i risultati"""
        results = {
            'database_size': [],
            'tablespace_size': [],
            'schema_size': [],
            'table_size': [],
            'index_size': [],
            'segment_size': [],
            'code_lines': [],
            'code_stats': [],
            'is_dba': is_dba
        }
        
        query_descriptions = {
            'database_size': 'Dimensioni database',
            'tablespace_size': 'Dimensioni tablespace',
            'schema_size': 'Dimensioni schema',
            'table_size': 'Dimensioni tabelle',
            'index_size': 'Dimensioni indici',
            'segment_size': 'Dimensioni segmenti',
            'code_lines': 'Righe codice',
            'code_stats': 'Statistiche codice'
        }
        
        for query_type, query_sql in queries.items():
            try:
                if query_type in ['database_size', 'tablespace_size'] and not is_dba:
                    # Query semplici per NON-DBA senza parametri
                    cursor.execute(query_sql)
                elif params and query_type in ['schema_size', 'table_size', 'index_size', 'segment_size', 'code_lines', 'code_stats']:
                    # Query che potrebbero aver bisogno di parametri per DBA
                    cursor.execute(query_sql, **params)
                else:
                    # Query senza parametri
                    cursor.execute(query_sql)
                
                results[query_type] = cursor.fetchall()
                count = len(results[query_type])
                print(f"    - {query_descriptions[query_type]}: {count} record")
                
            except Exception as e:
                print(f"    ⚠️  Errore query {query_descriptions[query_type]}: {str(e)}")
                results[query_type] = []
        
        cursor.close()
        return results
    
    def get_oracle_sizes(self, connection, is_dba, db_config):
        """🆕 Dispatcher per query dimensioni in base ai privilegi"""
        if is_dba:
            return self.get_oracle_sizes_dba(connection, db_config)
        else:
            return self.get_oracle_sizes_non_dba(connection)
    
    def get_all_schemas_for_dba(self, connection, db_config):
        """🆕 Ottiene lista di tutti gli schemi per utenti DBA (per ora2pg)"""
        cursor = connection.cursor()
        
        # Verifica se analizzare tutti gli schemi
        analyze_all = db_config.get('analyze_all_schemas', True)
        
        if not analyze_all:
            # Solo schema dell'utente corrente
            cursor.execute("SELECT USER FROM DUAL")
            current_user = cursor.fetchone()[0]
            cursor.close()
            return [current_user]
        
        try:
            # Per DBA: ottieni tutti gli schemi non di sistema con oggetti significativi
            cursor.execute("""
                SELECT DISTINCT owner
                FROM dba_objects
                WHERE owner NOT IN ('SYS', 'SYSTEM', 'PUBLIC', 'OUTLN', 'DBSNMP', 'ANONYMOUS', 'CTXSYS', 'DBSNMP', 'EXFSYS', 'LBACSYS', 'MDSYS', 'MGMT_VIEW', 'OLAPSYS', 'ORDDATA', 'OWBSYS', 'ORDPLUGINS', 'ORDSYS', 'PERFSTAT', 'WKPROXY', 'WKSYS', 'WK_TEST', 'WWWSCHEMAS', 'XDB', 'APEX_PUBLIC_USER', 'FLOWS_FILES')
                AND owner NOT LIKE '%$%'
                AND EXISTS (
                    SELECT 1 FROM dba_objects o2
                    WHERE o2.owner = dba_objects.owner
                    AND o2.object_type IN ('TABLE', 'VIEW', 'PROCEDURE', 'FUNCTION', 'PACKAGE', 'TRIGGER')
                )
                ORDER BY owner
            """)
            
            schemas = [row[0] for row in cursor.fetchall()]
            cursor.close()
            
            if not schemas:
                # Fallback: almeno lo schema dell'utente corrente
                cursor = connection.cursor()
                cursor.execute("SELECT USER FROM DUAL")
                current_user = cursor.fetchone()[0]
                cursor.close()
                return [current_user]
            
            return schemas
            
        except Exception as e:
            print(f"    ⚠️  Errore ottenimento schemi per DBA: {str(e)}")
            # Fallback: schema dell'utente corrente
            cursor.execute("SELECT USER FROM DUAL")
            current_user = cursor.fetchone()[0]
            cursor.close()
            return [current_user]
    
    def save_to_csv(self, data, filename, headers):
        """Salva dati in formato CSV con encoding UTF-8"""
        if not self.generate_csv:
            return  # CSV disabilitato
            
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                # Scrivi intestazioni
                writer.writerow(headers)
                # Scrivi dati
                writer.writerows(data)
            print(f"    > CSV salvato: {filename}")
        except Exception as e:
            print(f"    ⚠️  Errore salvataggio CSV {filename}: {str(e)}")
    
    def save_to_excel(self, data, filename, headers, sheet_name="Data"):
        """Salva dati in formato Excel con formattazione"""
        if not self.generate_excel:
            return  # Excel disabilitato
            
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # Crea workbook e worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Stili per le intestazioni
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Scrivi intestazioni con formattazione
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Scrivi dati
            for row_num, row_data in enumerate(data, 2):  # Inizia dalla riga 2
                for col_num, value in enumerate(row_data, 1):
                    # Gestione valori None/NULL
                    display_value = value if value is not None else ""
                    ws.cell(row=row_num, column=col_num, value=display_value)
            
            # Regola larghezza colonne automaticamente
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Imposta larghezza (con limite massimo)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Aggiungi filtri automatici
            if data:  # Solo se ci sono dati
                ws.auto_filter.ref = f"A1:{ws.cell(row=len(data)+1, column=len(headers)).coordinate}"
            
            # Salva file
            wb.save(filepath)
            print(f"    > Excel salvato: {filename}")
            
        except Exception as e:
            print(f"    ⚠️  Errore salvataggio Excel {filename}: {str(e)}")

    def save_combined_excel_report(self, oracle_data, connection_name, schema_name, is_dba):
        """🆕 Crea un file Excel completo con tutti i dati in fogli separati (include tipo DBA) - MODIFICATO NAMING"""
        if not self.generate_excel:
            return  # Excel disabilitato
            
        dba_suffix = "_dba" if is_dba else "_nodba"
        filename = f"{connection_name}_complete_analysis{dba_suffix}_{schema_name}.xlsx"  # 🔧 SCHEMA INVECE DI USERNAME
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            wb = openpyxl.Workbook()
            
            # Rimuovi il foglio di default
            wb.remove(wb.active)
            
            # Stili comuni
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 1. Foglio Sommario
            ws_summary = wb.create_sheet("Sommario")
            summary_data = [
                ["Tipo Analisi", "Numero Record"],
                ["Tipo Utente", "DBA" if is_dba else "NON-DBA"],
                ["--- DIPENDENZE ---", ""],
                ["Dipendenze", len(oracle_data.get('dependencies', []))],
                ["DB Links", len(oracle_data.get('db_links', []))],
                ["Privilegi Cross-Schema", len(oracle_data.get('cross_schema_privs', []))],
                ["Riferimenti Esterni", len(oracle_data.get('external_references', []))],
                ["Tipi Oggetti", len(oracle_data.get('object_summary', []))]
            ]
            
            # Aggiungi info dimensioni se disponibili
            if oracle_data.get('size_data'):
                size_data = oracle_data['size_data']
                summary_data.extend([
                    ["--- DIMENSIONI ---", ""],
                    ["Tabelle", len(size_data.get('table_size', []))],
                    ["Indici", len(size_data.get('index_size', []))],
                    ["Segmenti", len(size_data.get('segment_size', []))],
                    ["Oggetti Codice", len(size_data.get('code_stats', []))]
                ])
            
            for row_num, (title, count) in enumerate(summary_data, 1):
                ws_summary.cell(row=row_num, column=1, value=title)
                ws_summary.cell(row=row_num, column=2, value=count)
                
                if row_num == 1:  # Header
                    for col in [1, 2]:
                        cell = ws_summary.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
            
            # Regola larghezza colonne sommario
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 15
            
            # 2. Fogli per ogni tipo di dato - ✅ TUTTI I FOGLI ABILITATI NEL REPORT COMPLETO
            sheets_config = [
                # ✅ RIATTIVATO - Dependencies nel report completo
                ("Dipendenze", oracle_data.get('dependencies', []), 
                 ['SOURCE_OWNER', 'SOURCE_NAME', 'SOURCE_TYPE', 'TARGET_OWNER', 'TARGET_NAME', 'TARGET_TYPE', 'DB_LINK']),
                
                ("DB_Links", oracle_data.get('db_links', []), 
                 ['OWNER', 'DB_LINK', 'USERNAME', 'HOST']),
                
                # ✅ RIATTIVATO - Objects nel report completo
                ("Oggetti", oracle_data.get('object_summary', []), 
                 ['OWNER', 'OBJECT_TYPE', 'COUNT']),
                
                # ✅ RIATTIVATO - Cross Schema Privileges nel report completo  
                ("Privilegi_Cross_Schema", oracle_data.get('cross_schema_privs', []), 
                 ['GRANTOR', 'GRANTEE', 'TABLE_SCHEMA', 'TABLE_NAME', 'PRIVILEGE']),
                
                # ✅ RIATTIVATO - External References nel report completo
                ("Riferimenti_Esterni", oracle_data.get('external_references', []), 
                 ['SYNONYM_OWNER', 'SYNONYM_NAME', 'REFERENCED_OWNER', 'REFERENCED_OBJECT', 'DB_LINK'])
            ]
            
            for sheet_name, data, headers in sheets_config:
                if data:  # Solo se ci sono dati
                    ws = wb.create_sheet(sheet_name)
                    
                    # Scrivi intestazioni
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Scrivi dati
                    for row_num, row_data in enumerate(data, 2):
                        for col_num, value in enumerate(row_data, 1):
                            display_value = value if value is not None else ""
                            ws.cell(row=row_num, column=col_num, value=display_value)
                    
                    # Regola larghezza colonne
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Aggiungi filtri
                    ws.auto_filter.ref = f"A1:{ws.cell(row=len(data)+1, column=len(headers)).coordinate}"
            
            # Salva file
            wb.save(filepath)
            print(f"    > Excel completo salvato: {filename}")
          
        except Exception as e:
          print(f"    ⚠️  Errore salvataggio Excel completo: {str(e)}")
    
    def save_sizes_excel_report(self, size_data, connection_name, schema_name):
        """🆕 Crea un file Excel dedicato alle dimensioni con headers aggiornati per tablespace - MODIFICATO NAMING"""
        if not self.generate_excel or not self.analyze_sizes:
            return  # Excel o analisi dimensioni disabilitati
            
        dba_status = "dba" if size_data.get('is_dba') else "nodba"
        filename = f"{connection_name}_sizes_analysis_{dba_status}_{schema_name}.xlsx"  # 🔧 SCHEMA INVECE DI USERNAME
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            wb = openpyxl.Workbook()
            
            # Rimuovi il foglio di default
            wb.remove(wb.active)
            
            # Stili comuni
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 🆕 Configurazione fogli per dimensioni con headers aggiornati per tablespace
            if size_data.get('is_dba'):
                # Headers DBA per tablespace (dettagli completi)
                tablespace_headers = [
                    'TABLESPACE_NAME', 'STATUS', 'TYPE', 'ALLOCATED_GB', 'ALLOCATED_MB', 'ALLOCATED_BYTES',
                    'USED_GB', 'USED_MB', 'USED_BYTES', 'FREE_GB', 'FREE_MB', 'FREE_BYTES',
                    'PCT_USED', 'PCT_FREE', 'DATAFILE_COUNT', 'SEGMENT_COUNT'
                ]
            else:
                # Headers NON-DBA per tablespace (aggregato)
                tablespace_headers = [
                    'TABLESPACE_NAME', 'USED_GB', 'USED_MB', 'USED_BYTES', 'FILE_COUNT', 'STATUS'
                ]
            
            size_sheets_config = [
                ("Database_Size", size_data.get('database_size', []), 
                 ['METRIC_TYPE', 'OBJECT_NAME', 'SIZE_GB', 'SIZE_MB', 'SIZE_BYTES', 'FILE_COUNT']),
                ("Tablespace_Size", size_data.get('tablespace_size', []), 
                 tablespace_headers),
                ("Schema_Size", size_data.get('schema_size', []), 
                 ['OWNER', 'SIZE_GB', 'SIZE_MB', 'SIZE_BYTES', 'SEGMENT_COUNT']),
                ("Table_Size", size_data.get('table_size', []), 
                 ['OWNER', 'TABLE_NAME', 'SEGMENT_TYPE', 'TABLESPACE_NAME', 'SIZE_GB', 'SIZE_MB', 'SIZE_BYTES', 'BLOCKS', 'EXTENTS']),
                ("Index_Size", size_data.get('index_size', []), 
                 ['OWNER', 'INDEX_NAME', 'SEGMENT_TYPE', 'TABLESPACE_NAME', 'SIZE_GB', 'SIZE_MB', 'SIZE_BYTES', 'BLOCKS', 'EXTENTS']),
                ("Segment_Size", size_data.get('segment_size', []), 
                 ['OWNER', 'SEGMENT_NAME', 'SEGMENT_TYPE', 'TABLESPACE_NAME', 'SIZE_GB', 'SIZE_MB', 'SIZE_BYTES', 'BLOCKS', 'EXTENTS', 'INITIAL_EXTENT', 'NEXT_EXTENT', 'MAX_EXTENTS']),
                ("Code_Stats", size_data.get('code_stats', []), 
                 ['OWNER', 'OBJECT_NAME', 'OBJECT_TYPE', 'TOTAL_LINES', 'TOTAL_CHARS', 'TOTAL_BYTES', 'FIRST_LINE', 'LAST_LINE'])
            ]
            
            # Crea fogli solo se ci sono dati
            for sheet_name, data, headers in size_sheets_config:
                if data:  # Solo se ci sono dati
                    ws = wb.create_sheet(sheet_name)
                    
                    # Scrivi intestazioni
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Scrivi dati
                    for row_num, row_data in enumerate(data, 2):
                        for col_num, value in enumerate(row_data, 1):
                            display_value = value if value is not None else ""
                            ws.cell(row=row_num, column=col_num, value=display_value)
                    
                    # Regola larghezza colonne
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Aggiungi filtri
                    ws.auto_filter.ref = f"A1:{ws.cell(row=len(data)+1, column=len(headers)).coordinate}"
            
            # Aggiungi foglio sommario dimensioni se ci sono dati
            if any(size_data.get(key, []) for key in ['database_size', 'tablespace_size', 'schema_size', 'table_size', 'index_size', 'segment_size', 'code_stats']):
                ws_summary = wb.create_sheet("Sommario_Dimensioni", 0)  # Inserisci come primo foglio
                
                summary_data = [["Categoria", "Numero Elementi", "Tipo Utente", "Note"]]
                user_type = "DBA" if size_data.get('is_dba') else "NON-DBA"
                
                if size_data.get('database_size'):
                    note = "Dimensioni totali database" if size_data.get('is_dba') else "Solo schema utente"
                    summary_data.append(["Database Size", len(size_data['database_size']), user_type, note])
                if size_data.get('tablespace_size'):
                    note = "Dettagli completi tablespace" if size_data.get('is_dba') else "Aggregato per tablespace utente"
                    summary_data.append(["Tablespace Size", len(size_data['tablespace_size']), user_type, note])
                if size_data.get('schema_size'):
                    note = "Tutti gli schemi" if size_data.get('is_dba') else "Solo schema utente"
                    summary_data.append(["Schema Size", len(size_data['schema_size']), user_type, note])
                if size_data.get('table_size'):
                    note = "Tutti gli schemi" if size_data.get('is_dba') else "Solo schema utente"
                    summary_data.append(["Table Size", len(size_data['table_size']), user_type, note])
                if size_data.get('index_size'):
                    note = "Tutti gli schemi" if size_data.get('is_dba') else "Solo schema utente"
                    summary_data.append(["Index Size", len(size_data['index_size']), user_type, note])
                if size_data.get('segment_size'):
                    note = "Tutti gli schemi" if size_data.get('is_dba') else "Solo schema utente"
                    summary_data.append(["Segment Size", len(size_data['segment_size']), user_type, note])
                if size_data.get('code_stats'):
                    note = "Tutti gli schemi accessibili" if size_data.get('is_dba') else "Solo schema utente"
                    summary_data.append(["Code Objects", len(size_data['code_stats']), user_type, note])
                
                # Scrivi dati sommario
                for row_num, row_data in enumerate(summary_data, 1):
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                        if row_num == 1:  # Header
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                
                # Regola larghezza colonne sommario
                ws_summary.column_dimensions['A'].width = 20
                ws_summary.column_dimensions['B'].width = 15
                ws_summary.column_dimensions['C'].width = 12
                ws_summary.column_dimensions['D'].width = 30
            
            # Salva file solo se ci sono fogli con dati
            if wb.worksheets:
                wb.save(filepath)
                print(f"    > Excel dimensioni salvato: {filename}")
            else:
                print(f"    ⚠️  Nessun dato dimensioni disponibile per Excel")
            
        except Exception as e:
            print(f"    ⚠️  Errore salvataggio Excel dimensioni: {str(e)}")
    
    def save_summary_report(self, all_results):
        """🆕 Crea un report riassuntivo in formato testo con info DBA/NON-DBA"""
        report_path = os.path.join(self.output_dir, 'summary_report.txt')
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("ORACLE MULTI-DATABASE DEPENDENCY ANALYSIS REPORT\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"Report generato: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Numero database analizzati: {len(all_results)}\n")
            f.write(f"Formato output: {'Excel' if self.generate_excel else 'CSV' if self.generate_csv else 'Solo Database'}\n")
            f.write(f"Output ora2pg: {self.ora2pg_output_mode}\n")
            f.write(f"Analisi dimensioni: {'Abilitata' if self.analyze_sizes else 'Disabilitata'}\n")
            f.write(f"Prefissi tabelle: pdt_dep_dba_/pdt_dep_nodba_ (dipendenze), pdt_sizes_dba_/pdt_sizes_nodba_ (dimensioni), ptd_ (ora2pg)\n")
            f.write(f"ORA2PG: Schema configurabile per connessione, fallback logica DBA/NON-DBA\n")
            f.write(f"Query tablespace: DBA (dettagli completi), NON-DBA (aggregato per tablespace)\n")
            f.write(f"MODIFICHE: File Excel specifici disattivati, naming con schema invece di username\n\n")
            
            total_deps = 0
            total_links = 0
            total_cross_schema = 0
            total_ext_refs = 0
            total_cost = 0.0
            total_schema_size_gb = 0.0
            total_tables = 0
            total_indexes = 0
            total_dba_users = 0
            total_non_dba_users = 0
            
            for db_key, results in all_results.items():
                f.write(f"\nDatabase: {db_key}\n")
                f.write("-" * 50 + "\n")
                
                if results.get('error'):
                    f.write(f"ERRORE: {results['error']}\n")
                    continue
                
                # Info privilegi utente
                is_dba = results.get('is_dba', False)
                if is_dba:
                    total_dba_users += 1
                else:
                    total_non_dba_users += 1
                
                f.write(f"Privilegi utente: {'DBA' if is_dba else 'NON-DBA'}\n")
                
                # Sommario oggetti
                f.write("\nOggetti database:\n")
                for obj in results.get('object_summary', []):
                    f.write(f"  - {obj[1]}: {obj[2]}\n")
                
                # Conteggi dipendenze
                deps = len(results.get('dependencies', []))
                links = len(results.get('db_links', []))
                cross = len(results.get('cross_schema_privs', []))
                ext_refs = len(results.get('external_references', []))
                
                f.write(f"\nDipendenze (modalità {'DBA' if is_dba else 'NON-DBA'}):\n")
                f.write(f"  - Dipendenze trovate: {deps}\n")
                f.write(f"  - DB Links: {links}\n")
                f.write(f"  - Privilegi cross-schema: {cross}\n")
                f.write(f"  - Riferimenti esterni: {ext_refs}\n")
                
                # Informazioni dimensioni
                if results.get('size_data'):
                    size_data = results['size_data']
                    f.write(f"\nDimensioni (modalità {'DBA' if is_dba else 'NON-DBA'}):\n")
                    
                    if size_data.get('schema_size'):
                        for schema in size_data['schema_size']:
                            f.write(f"  - Schema {schema[0]}: {schema[1]} GB\n")
                            total_schema_size_gb += float(schema[1]) if schema[1] else 0
                    
                    tables_count = len(size_data.get('table_size', []))
                    indexes_count = len(size_data.get('index_size', []))
                    f.write(f"  - Tabelle: {tables_count}\n")
                    f.write(f"  - Indici: {indexes_count}\n")
                    total_tables += tables_count
                    total_indexes += indexes_count
                    
                    if size_data.get('code_stats'):
                        total_objects = len(size_data['code_stats'])
                        total_lines = sum(obj[3] for obj in size_data['code_stats'] if obj[3])
                        total_bytes = sum(obj[5] for obj in size_data['code_stats'] if obj[5])
                        f.write(f"  - Oggetti codice: {total_objects}\n")
                        f.write(f"  - Righe codice totali: {total_lines}\n")
                        f.write(f"  - Bytes codice totali: {total_bytes}\n")
                
                # Metriche ora2pg
                if 'ora2pg_metrics' in results:
                    cost = results['ora2pg_metrics'].get('total_cost', 0)
                    level = results['ora2pg_metrics'].get('migration_level', 'Unknown')
                    analyzed_schemas = results['ora2pg_metrics'].get('analyzed_schemas', [])
                    target_schema = results['ora2pg_metrics'].get('target_schema', 'auto')
                    f.write(f"\nStima migrazione ora2pg:\n")
                    f.write(f"  - Schema target: {target_schema}\n")
                    f.write(f"  - Costo totale: {cost}\n")
                    f.write(f"  - Livello: {level}\n")
                    if analyzed_schemas:
                        f.write(f"  - Schemi analizzati: {', '.join(analyzed_schemas)}\n")
                    total_cost += float(cost) if cost else 0
                
                total_deps += deps
                total_links += links
                total_cross_schema += cross
                total_ext_refs += ext_refs
            
            # Sommario totale
            f.write(f"\n{'='*60}\n")
            f.write("SOMMARIO TOTALE\n")
            f.write(f"{'='*60}\n")
            f.write(f"Dipendenze totali: {total_deps}\n")
            f.write(f"DB Links totali: {total_links}\n")
            f.write(f"Privilegi cross-schema totali: {total_cross_schema}\n")
            f.write(f"Riferimenti esterni totali: {total_ext_refs}\n")
            f.write(f"Costo totale stimato migrazione: {total_cost}\n")
            
            if self.analyze_sizes:
                f.write(f"Dimensioni totali schemi: {total_schema_size_gb:.2f} GB\n")
                f.write(f"Tabelle totali: {total_tables}\n")
                f.write(f"Indici totali: {total_indexes}\n")
            
            f.write(f"Utenti DBA: {total_dba_users}\n")
            f.write(f"Utenti NON-DBA: {total_non_dba_users}\n")
            f.write(f"ORA2PG: Schema configurabile per connessione, fallback logica DBA/NON-DBA\n")
            f.write(f"Query tablespace: DBA (dettagli completi), NON-DBA (aggregato per tablespace)\n")
            f.write(f"MODIFICHE: File Excel specifici disattivati, naming con schema invece di username\n")
        
        print(f"\n> Report riassuntivo salvato: summary_report.txt")
    
    def run_ora2pg_analysis(self, dsn, username, password, connection_name, is_dba, db_config):
        """🆕 Esegue ora2pg con schema configurabile o logica DBA/NON-DBA - MODIFICATO NAMING"""
        # Parsing DSN per ora2pg
        import re
        dsn_pattern = r'([^:]+):(\d+)/(.+)'
        match = re.match(dsn_pattern, dsn)
        
        if match:
            host = match.group(1)
            port = match.group(2)
            service = match.group(3)
            oracle_dsn = f"//{host}:{port}/{service}"
        else:
            oracle_dsn = dsn
        
        # 🆕 GESTIONE SCHEMA CONFIGURABILE
        target_schema = None
        
        # 🔧 DETERMINA SCHEMA NAME PER NAMING
        schema_name = db_config.get('schema', username)
        
        if 'schema' in db_config and db_config['schema']:
            # Schema specificato nella configurazione
            target_schema = db_config['schema']
            analyzed_schemas = [target_schema]
            print(f"    🎯 Schema configurato: {target_schema}")
            print(f"    📋 ORA2PG analizzerà lo schema: {target_schema}")
        else:
            # Fallback alla logica DBA/NON-DBA
            if is_dba:
                print(f"    🔧 Modalità DBA: preparazione analisi completa database...")
                
                # Connetti temporaneamente per ottenere lista schemi
                try:
                    temp_connection = self.get_db_connection(dsn, username, password)
                    analyzed_schemas = self.get_all_schemas_for_dba(temp_connection, db_config)
                    temp_connection.close()
                    
                    print(f"    📋 Schemi da analizzare per utente DBA: {len(analyzed_schemas)}")
                    for schema in analyzed_schemas[:10]:  # Mostra i primi 10
                        print(f"      - {schema}")
                    if len(analyzed_schemas) > 10:
                        print(f"      ... e altri {len(analyzed_schemas) - 10} schemi")
                        
                except Exception as e:
                    print(f"    ⚠️  Errore ottenimento schemi per DBA: {str(e)}")
                    print(f"    🔄 Fallback: analisi solo schema {username}")
                    analyzed_schemas = [username]
            else:
                # NON-DBA: solo schema dell'utente
                analyzed_schemas = [username]
                print(f"    🔧 Modalità NON-DBA: analisi solo schema {username}")
        
        # 🔧 DETERMINA MODALITÀ PER IL NOME FILE - MODIFICATO
        if target_schema:
            mode_desc = f"CONFIGURED"  # 🔧 SOLO "CONFIGURED" per file configurati
            analysis_mode = f"Schema configurato: {target_schema}"
            final_name = schema_name  # 🔧 SOLO SCHEMA PER FILE CONFIGURED
        elif is_dba and len(analyzed_schemas) > 1:
            mode_desc = "DBA_FULL"
            analysis_mode = f"DBA - tutti gli schemi ({len(analyzed_schemas)})"
            final_name = f"{mode_desc}_{schema_name}"
        else:
            mode_desc = "SINGLE_SCHEMA"
            analysis_mode = f"Schema singolo: {analyzed_schemas[0]}"
            final_name = f"{mode_desc}_{schema_name}"
        
        print(f"    🚀 Esecuzione ora2pg - {analysis_mode}")
        
        # Crea configurazione ora2pg
        if target_schema:
            # Schema configurato
            ora2pg_conf_content = f"""# Ora2pg configuration file for migration assessment
# Connection: {connection_name} - Schema Configured: {target_schema}

# Oracle database connection
ORACLE_DSN      dbi:Oracle:{oracle_dsn}
ORACLE_USER     {username}
ORACLE_PWD      {password}

# PostgreSQL target version
PG_VERSION      14

# Output settings
OUTPUT_DIR      {self.output_dir}
TYPE            SHOW_REPORT
ESTIMATE_COST   1

# Report settings
TOP_MAX         50
HUMAN_DAYS_LIMIT 5
COST_UNIT_VALUE 500

# Configured schema
SCHEMA          {target_schema}

# Character set
NLS_LANG        AMERICAN_AMERICA.AL32UTF8
BINMODE         utf8

# Disable some features for assessment
SKIP_INDEXES    0
SKIP_CONSTRAINTS 0
SKIP_TRIGGERS   0

# Debug
DEBUG           0
"""
        elif is_dba and len(analyzed_schemas) > 1:
            # DBA con più schemi
            ora2pg_conf_content = f"""# Ora2pg configuration file for migration assessment - DBA MODE
# Connection: {connection_name} - DBA Full Database Analysis
# Analyzed schemas: {', '.join(analyzed_schemas)}

# Oracle database connection
ORACLE_DSN      dbi:Oracle:{oracle_dsn}
ORACLE_USER     {username}
ORACLE_PWD      {password}

# PostgreSQL target version
PG_VERSION      14

# Output settings
OUTPUT_DIR      {self.output_dir}
TYPE            SHOW_REPORT
ESTIMATE_COST   1

# Report settings
TOP_MAX         50
HUMAN_DAYS_LIMIT 5
COST_UNIT_VALUE 500

# DBA MODE: Multiple schemas analysis
SCHEMA          {','.join(analyzed_schemas) if len(analyzed_schemas) <= 10 else f'{username},+ALL_SCHEMA'}

# Character set
NLS_LANG        AMERICAN_AMERICA.AL32UTF8
BINMODE         utf8

# DBA specific settings
SKIP_FKEYS      0
SKIP_INDEXES    0
SKIP_CONSTRAINTS 0
SKIP_TRIGGERS   0
SKIP_CHECKS     0

# Debug
DEBUG           0
"""
        else:
            # Schema singolo (NON-DBA o DBA con un solo schema)
            ora2pg_conf_content = f"""# Ora2pg configuration file for migration assessment
# Connection: {connection_name} - Single Schema: {analyzed_schemas[0]}

# Oracle database connection
ORACLE_DSN      dbi:Oracle:{oracle_dsn}
ORACLE_USER     {username}
ORACLE_PWD      {password}

# PostgreSQL target version
PG_VERSION      14

# Output settings
OUTPUT_DIR      {self.output_dir}
TYPE            SHOW_REPORT
ESTIMATE_COST   1

# Report settings
TOP_MAX         50
HUMAN_DAYS_LIMIT 5
COST_UNIT_VALUE 500

# Schema to analyze
SCHEMA          {analyzed_schemas[0]}

# Character set
NLS_LANG        AMERICAN_AMERICA.AL32UTF8
BINMODE         utf8

# Disable some features for assessment
SKIP_INDEXES    0
SKIP_CONSTRAINTS 0
SKIP_TRIGGERS   0

# Debug
DEBUG           0
"""
        
        # 🔧 File ora2pg con naming aggiornato
        conf_file = os.path.join(self.output_dir, f'{connection_name}_ora2pg_{final_name}.conf')
        with open(conf_file, 'w', encoding='utf-8') as f:
            f.write(ora2pg_conf_content)
        
        print(f"    > Config ora2pg creato: {connection_name}_ora2pg_{final_name}.conf")
        
        # 🔧 Report ora2pg con naming aggiornato
        html_output_file = os.path.join(self.output_dir, f'{connection_name}_migration_report_{final_name}.html')
        txt_output_file = os.path.join(self.output_dir, f'{connection_name}_migration_report_{final_name}.txt')
        
        # Esegui ora2pg
        report_result = self._execute_ora2pg_command(conf_file, html_output_file, txt_output_file, connection_name, final_name)
        
        if report_result:
            # Aggiungi info configurazione ai risultati
            report_result['analyzed_schemas'] = analyzed_schemas
            report_result['dba_mode'] = is_dba and not target_schema  # Solo se è DBA e non ha schema configurato
            report_result['reports_count'] = 1
            report_result['target_schema'] = target_schema if target_schema else 'auto'
            
            print(f"    ✅ Analisi ora2pg completata - {analysis_mode}")
            print(f"      - Costo totale: {report_result.get('total_cost', 'N/A')}")
            print(f"      - Oggetti trovati: {len(report_result.get('ora2pg_object_summary', []))}")
            
            return report_result
        else:
            print(f"    ⚠️  Errore ora2pg per {connection_name}")
            return None
    
    def _execute_ora2pg_command(self, conf_file, html_output_file, txt_output_file, connection_name, final_name):
        """🆕 Esegue il comando ora2pg e gestisce l'output configurabile - MODIFICATO NAMING"""
        try:
            # 🆕 GESTIONE OUTPUT CONFIGURABILE
            if self.ora2pg_output_mode == 'html_only':
                # Solo output HTML
                if sys.platform == 'win32':
                    cmd = f'ora2pg -c "{conf_file}" --type=SHOW_REPORT --estimate_cost --dump_as_html > "{html_output_file}"'
                else:
                    cmd = f'ora2pg -c "{conf_file}" --type=SHOW_REPORT --estimate_cost --dump_as_html > "{html_output_file}"'
                
                result = subprocess.run(cmd, shell=True, capture_output=False, text=True)
                
                if result.returncode == 0:
                    print(f"    > Report ora2pg HTML generato: {os.path.basename(html_output_file)}")
                    return self.parse_ora2pg_report(html_output_file, None)
                else:
                    print(f"    ⚠️  Errore ora2pg per {final_name}")
                    return None
                    
            elif self.ora2pg_output_mode == 'html_and_txt':
                # Output HTML + TXT
                # Prima genera HTML
                if sys.platform == 'win32':
                    cmd_html = f'ora2pg -c "{conf_file}" --type=SHOW_REPORT --estimate_cost --dump_as_html > "{html_output_file}"'
                else:
                    cmd_html = f'ora2pg -c "{conf_file}" --type=SHOW_REPORT --estimate_cost --dump_as_html > "{html_output_file}"'
                
                result_html = subprocess.run(cmd_html, shell=True, capture_output=False, text=True)
                
                # Poi genera TXT
                if sys.platform == 'win32':
                    cmd_txt = f'ora2pg -c "{conf_file}" --type=SHOW_REPORT --estimate_cost > "{txt_output_file}"'
                else:
                    cmd_txt = f'ora2pg -c "{conf_file}" --type=SHOW_REPORT --estimate_cost > "{txt_output_file}"'
                    
                result_txt = subprocess.run(cmd_txt, shell=True, capture_output=False, text=True)
                
                if result_html.returncode == 0:
                    print(f"    > Report ora2pg HTML generato: {os.path.basename(html_output_file)}")
                    
                if result_txt.returncode == 0:
                    print(f"    > Report ora2pg TXT generato: {os.path.basename(txt_output_file)}")
                
                if result_html.returncode == 0:
                    return self.parse_ora2pg_report(html_output_file, txt_output_file if result_txt.returncode == 0 else None)
                else:
                    print(f"    ⚠️  Errore ora2pg per {final_name}")
                    return None
            else:
                print(f"    ⚠️  Modalità output ora2pg non riconosciuta: {self.ora2pg_output_mode}")
                return None
                
        except Exception as e:
            print(f"    ⚠️  Errore esecuzione ora2pg per {final_name}: {str(e)}")
            return None
    
    def parse_ora2pg_report(self, html_file, txt_file):
        """Parsing del report ora2pg per estrarre metriche"""
        metrics = {
            'total_cost': 0,
            'migration_level': 'Unknown',
            'objects_count': {},
            'details': [],
            'ora2pg_object_summary': []
        }
        
        # Parsing del file testuale per metriche più precise (se disponibile)
        if txt_file and os.path.exists(txt_file):
            try:
                with open(txt_file, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                
                # Pattern per estrarre costo totale
                cost_patterns = [
                    r'Total\s+estimated\s+cost:\s*(\d+\.?\d*)',
                    r'Total\s+cost:\s*(\d+\.?\d*)',
                    r'Migration\s+cost:\s*(\d+\.?\d*)'
                ]
                
                for pattern in cost_patterns:
                    cost_match = re.search(pattern, content, re.IGNORECASE)
                    if cost_match:
                        metrics['total_cost'] = float(cost_match.group(1))
                        break
                
                # Pattern per livello di migrazione
                level_patterns = [
                    r'Migration\s+level:\s*(\w+)',
                    r'Level:\s*(\w+)'
                ]
                
                for pattern in level_patterns:
                    level_match = re.search(pattern, content, re.IGNORECASE)
                    if level_match:
                        metrics['migration_level'] = level_match.group(1)
                        break
                
                # Estrai conteggio oggetti
                object_pattern = r'(\w+)\s+\[(\d+)\]'
                for match in re.finditer(object_pattern, content):
                    obj_type = match.group(1)
                    obj_count = int(match.group(2))
                    metrics['objects_count'][obj_type] = obj_count
                    
            except Exception as e:
                print(f"    ⚠️  Errore parsing report TXT: {str(e)}")
        
        # Parsing del file HTML per object summary (sempre disponibile)
        if html_file and os.path.exists(html_file):
            try:
                object_summary = self.parse_object_summary_from_html(html_file)
                metrics['ora2pg_object_summary'] = object_summary
                print(f"    > Parsed {len(object_summary)} oggetti dal report HTML")
            except Exception as e:
                print(f"    ⚠️  Errore parsing HTML: {str(e)}")
        
        return metrics
    
    def parse_object_summary_from_html(self, html_path):
        """Parse della tabella Object Summary dal report HTML ora2pg con dettagli espansi"""
        with open(html_path, 'r', encoding='utf-8', errors='ignore') as f:
            soup = BeautifulSoup(f, 'html.parser')

        summary = []
        tables = soup.find_all('table')

        for table in tables:
            headers = [th.text.strip().lower() for th in table.find_all('th')]
            # Cerca la tabella con le colonne corrette (case-insensitive)
            if len(headers) >= 6 and 'object' in headers[0] and 'number' in headers[1]:
                print(f"    > Trovata tabella Object Summary con headers: {headers}")
                for row in table.find_all('tr')[1:]:  # Skip header row
                    cols = row.find_all('td')
                    if len(cols) < 6:
                        continue
                    try:
                        # Gestisce valori numerici che potrebbero essere vuoti o non numerici
                        def safe_int(value):
                            try:
                                return int(value.strip()) if value.strip() else 0
                            except:
                                return 0
                        
                        def safe_float(value):
                            try:
                                return float(value.strip()) if value.strip() else 0.0
                            except:
                                return 0.0
                        
                        object_name = cols[0].text.strip()
                        object_number = safe_int(cols[1].text)
                        invalid_count = safe_int(cols[2].text)
                        estimated_cost = safe_float(cols[3].text)
                        comments = cols[4].text.strip()
                        details = cols[5].text.strip()
                        
                        # Aggiungi il record principale
                        summary.append({
                            'object_name': object_name,
                            'object_number': object_number,
                            'invalid_count': invalid_count,
                            'estimated_cost': estimated_cost,
                            'comments': comments,
                            'details': details,
                            'detail_type': 'MAIN',
                            'procedure_name': None,
                            'procedure_cost': None
                        })
                        
                        # Parse dei dettagli per estrarre procedure/funzioni individuali
                        if details and details != '-':
                            parsed_details = self.parse_procedure_details(details)
                            for proc_detail in parsed_details:
                                summary.append({
                                    'object_name': object_name,
                                    'object_number': 1,  # Ogni procedura conta come 1
                                    'invalid_count': 0,
                                    'estimated_cost': proc_detail['cost'],
                                    'comments': f"Detail of {object_name}",
                                    'details': proc_detail['full_name'],
                                    'detail_type': 'PROCEDURE',
                                    'procedure_name': proc_detail['name'],
                                    'procedure_cost': proc_detail['cost']
                                })
                        
                    except Exception as e:
                        print(f"    ⚠️  Errore parsing riga tabella: {e}")
                        continue
                break  # Found the right table, stop searching

        print(f"    > Oggetti trovati nella tabella (inclusi dettagli): {len(summary)}")
        return summary

    def parse_procedure_details(self, details_text):
        """Parse del testo dei dettagli per estrarre le singole procedure/funzioni"""
        procedures = []
        
        # Pattern per trovare: nome_procedure: costo
        # Esempi: "pck_t_user.validaupdate: 3", "package_analyze.estrai_nome_procedura: 4"
        pattern = r'([a-zA-Z_][a-zA-Z0-9_.]*)\s*:\s*(\d+(?:\.\d+)?)'
        
        matches = re.findall(pattern, details_text)
        
        for match in matches:
            proc_name = match[0].strip()
            try:
                proc_cost = float(match[1])
                procedures.append({
                    'name': proc_name,
                    'cost': proc_cost,
                    'full_name': f"{proc_name}: {proc_cost}"
                })
            except ValueError:
                print(f"    ⚠️  Errore parsing costo per {proc_name}: {match[1]}")
                continue
        
        if procedures:
            print(f"      > Estratte {len(procedures)} procedure/funzioni dai dettagli")
        return procedures
    
    def create_database_schema(self):
        """🆕 Crea lo schema del database con prefissi pdt_dep_dba_/pdt_dep_nodba_, pdt_sizes_dba_/pdt_sizes_nodba_ e FK"""
        try:
            conn = psycopg2.connect(**self.pg_config)
            cursor = conn.cursor()
            
            # Crea schema dedicato
            cursor.execute("CREATE SCHEMA IF NOT EXISTS oracle_migration")
            
            # 🆕 TABELLE CON NUOVI PREFISSI E CAMPO SCHEMA
            # 🔧 Crea tabella pdt_connections con campo schema aggiunto
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_connections (
                    id SERIAL PRIMARY KEY,
                    connection_name VARCHAR(100) UNIQUE NOT NULL,
                    dsn VARCHAR(255) NOT NULL,
                    username VARCHAR(100) NOT NULL,
                    schema VARCHAR(100),
                    description TEXT,
                    is_dba BOOLEAN DEFAULT FALSE,
                    analyze_all_schemas BOOLEAN DEFAULT FALSE,
                    created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            
            
            # 🆕 TABELLE DIPENDENZE DBA CON PREFISSO pdt_dep_dba_
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_dep_dba_dependencies (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    source_owner VARCHAR(100),
                    source_name VARCHAR(255),
                    source_type VARCHAR(50),
                    target_owner VARCHAR(100),
                    target_name VARCHAR(255),
                    target_type VARCHAR(50),
                    db_link VARCHAR(100)
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_dep_dba_db_links (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    db_link VARCHAR(255),
                    username VARCHAR(100),
                    host VARCHAR(255)
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_dep_dba_cross_schema_privileges (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    grantor VARCHAR(100),
                    grantee VARCHAR(100),
                    table_schema VARCHAR(100),
                    table_name VARCHAR(255),
                    privilege VARCHAR(50)
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_dep_dba_external_references (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    synonym_owner VARCHAR(100),
                    synonym_name VARCHAR(255),
                    referenced_owner VARCHAR(100),
                    referenced_object VARCHAR(255),
                    db_link VARCHAR(100)
                )
            """)
            
            # 🆕 TABELLE DIPENDENZE NON-DBA CON PREFISSO pdt_dep_nodba_
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_dep_nodba_dependencies (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    source_owner VARCHAR(100),
                    source_name VARCHAR(255),
                    source_type VARCHAR(50),
                    target_owner VARCHAR(100),
                    target_name VARCHAR(255),
                    target_type VARCHAR(50),
                    db_link VARCHAR(100)
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_dep_nodba_db_links (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    db_link VARCHAR(255),
                    username VARCHAR(100),
                    host VARCHAR(255)
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_dep_nodba_cross_schema_privileges (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    grantor VARCHAR(100),
                    grantee VARCHAR(100),
                    table_schema VARCHAR(100),
                    table_name VARCHAR(255),
                    privilege VARCHAR(50)
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_dep_nodba_external_references (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    synonym_owner VARCHAR(100),
                    synonym_name VARCHAR(255),
                    referenced_owner VARCHAR(100),
                    referenced_object VARCHAR(255),
                    db_link VARCHAR(100)
                )
            """)
            
            # 🆕 TABELLE ORA2PG CON PREFISSO ptd_ (aggiornate con campi schema)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.ptd_ora2pg_estimates (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    schema_name VARCHAR(100),
                    total_cost NUMERIC(10,2),
                    migration_level VARCHAR(50),
                    analyzed_schemas TEXT,
                    target_schema VARCHAR(100),
                    dba_mode BOOLEAN DEFAULT FALSE,
                    reports_count INTEGER DEFAULT 1,
                    metrics JSONB
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.ptd_ora2pg_object_summary (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    schema_name VARCHAR(100),
                    object_name TEXT,
                    object_number INTEGER,
                    invalid_count INTEGER,
                    estimated_cost NUMERIC(10,2),
                    comments TEXT,
                    details TEXT,
                    detail_type VARCHAR(20) DEFAULT 'MAIN',
                    procedure_name TEXT,
                    procedure_cost NUMERIC(10,2)
                )
            """)
            
            # 🆕 TABELLE DIMENSIONI DBA CON PREFISSO pdt_sizes_dba_ E COLONNE TABLESPACE AGGIORNATE
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_dba_database_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    metric_type VARCHAR(50),
                    object_name VARCHAR(100),
                    size_gb NUMERIC(12,2),
                    size_mb NUMERIC(12,2),
                    size_bytes BIGINT,
                    file_count INTEGER
                )
            """)
            
            # 🔧 Tablespace DBA con colonne aggiornate per query dettagliata
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_dba_tablespace_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    tablespace_name VARCHAR(100),
                    status VARCHAR(20),
                    type VARCHAR(20),
                    allocated_gb NUMERIC(12,2),
                    allocated_mb NUMERIC(12,2),
                    allocated_bytes BIGINT,
                    used_gb NUMERIC(12,2),
                    used_mb NUMERIC(12,2),
                    used_bytes BIGINT,
                    free_gb NUMERIC(12,2),
                    free_mb NUMERIC(12,2),
                    free_bytes BIGINT,
                    pct_used NUMERIC(5,2),
                    pct_free NUMERIC(5,2),
                    datafile_count INTEGER,
                    segment_count INTEGER
                )
            """)
            
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_dba_schema_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    size_gb NUMERIC(12,2),
                    size_mb NUMERIC(12,2),
                    size_bytes NUMERIC,
                    segment_count INTEGER
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_dba_table_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    table_name VARCHAR(255),
                    segment_type VARCHAR(50),
                    tablespace_name VARCHAR(100),
                    size_gb NUMERIC(12,4),
                    size_mb NUMERIC(12,2),
                    size_bytes BIGINT,
                    blocks BIGINT,
                    extents INTEGER
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_dba_index_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    index_name VARCHAR(255),
                    segment_type VARCHAR(50),
                    tablespace_name VARCHAR(100),
                    size_gb NUMERIC(12,4),
                    size_mb NUMERIC(12,2),
                    size_bytes BIGINT,
                    blocks BIGINT,
                    extents INTEGER
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_dba_segment_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    segment_name VARCHAR(255),
                    segment_type VARCHAR(50),
                    tablespace_name VARCHAR(100),
                    size_gb NUMERIC(12,4),
                    size_mb NUMERIC(12,2),
                    size_bytes BIGINT,
                    blocks BIGINT,
                    extents INTEGER,
                    initial_extent BIGINT,
                    next_extent BIGINT,
                    max_extents BIGINT
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_dba_code_lines (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    object_name VARCHAR(255),
                    object_type VARCHAR(50),
                    char_length INTEGER,
                    byte_length INTEGER,
                    line_number INTEGER,
                    line_text TEXT
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_dba_code_stats (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    object_name VARCHAR(255),
                    object_type VARCHAR(50),
                    total_lines INTEGER,
                    total_chars BIGINT,
                    total_bytes BIGINT,
                    first_line INTEGER,
                    last_line INTEGER
                )
            """)
            
            # 🆕 TABELLE DIMENSIONI NON-DBA CON PREFISSO pdt_sizes_nodba_ E COLONNE TABLESPACE AGGIORNATE
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_nodba_database_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    metric_type VARCHAR(50),
                    object_name VARCHAR(100),
                    size_gb NUMERIC(12,2),
                    size_mb NUMERIC(12,2),
                    size_bytes BIGINT,
                    file_count INTEGER
                )
            """)
            
            # 🔧 Tablespace NON-DBA con colonne aggiornate per query aggregata
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_nodba_tablespace_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    tablespace_name VARCHAR(100),
                    used_gb NUMERIC(12,2),
                    used_mb NUMERIC(12,2),
                    used_bytes BIGINT,
                    file_count INTEGER,
                    status VARCHAR(20)
                )
            """)
            
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_nodba_schema_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    size_gb NUMERIC(12,2),
                    size_mb NUMERIC(12,2),
                    size_bytes BIGINT,
                    segment_count INTEGER
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_nodba_table_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    table_name VARCHAR(255),
                    segment_type VARCHAR(50),
                    tablespace_name VARCHAR(100),
                    size_gb NUMERIC(12,4),
                    size_mb NUMERIC(12,2),
                    size_bytes BIGINT,
                    blocks BIGINT,
                    extents INTEGER
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_nodba_index_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    index_name VARCHAR(255),
                    segment_type VARCHAR(50),
                    tablespace_name VARCHAR(100),
                    size_gb NUMERIC(12,4),
                    size_mb NUMERIC(12,2),
                    size_bytes BIGINT,
                    blocks BIGINT,
                    extents INTEGER
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_nodba_segment_size (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    segment_name VARCHAR(255),
                    segment_type VARCHAR(50),
                    tablespace_name VARCHAR(100),
                    size_gb NUMERIC(12,4),
                    size_mb NUMERIC(12,2),
                    size_bytes BIGINT,
                    blocks BIGINT,
                    extents INTEGER,
                    initial_extent BIGINT,
                    next_extent BIGINT,
                    max_extents BIGINT
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_nodba_code_lines (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    object_name VARCHAR(255),
                    object_type VARCHAR(50),
                    char_length INTEGER,
                    byte_length INTEGER,
                    line_number INTEGER,
                    line_text TEXT
                )
            """)
            
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS oracle_migration.pdt_sizes_nodba_code_stats (
                    id SERIAL PRIMARY KEY,
                    connection_id INTEGER REFERENCES oracle_migration.pdt_connections(id) ON DELETE CASCADE,
                    analysis_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    owner VARCHAR(100),
                    object_name VARCHAR(255),
                    object_type VARCHAR(50),
                    total_lines INTEGER,
                    total_chars BIGINT,
                    total_bytes BIGINT,
                    first_line INTEGER,
                    last_line INTEGER
                )
            """)
            
            # 🆕 CREA INDICI CON NUOVI PREFISSI
            # Indici dipendenze DBA pdt_dep_dba_
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_dep_dba_dependencies_connection_id ON oracle_migration.pdt_dep_dba_dependencies(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_dep_dba_db_links_connection_id ON oracle_migration.pdt_dep_dba_db_links(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_dep_dba_cross_schema_privileges_connection_id ON oracle_migration.pdt_dep_dba_cross_schema_privileges(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_dep_dba_external_references_connection_id ON oracle_migration.pdt_dep_dba_external_references(connection_id)")
            
            # Indici dipendenze NON-DBA pdt_dep_nodba_
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_dep_nodba_dependencies_connection_id ON oracle_migration.pdt_dep_nodba_dependencies(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_dep_nodba_db_links_connection_id ON oracle_migration.pdt_dep_nodba_db_links(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_dep_nodba_cross_schema_privileges_connection_id ON oracle_migration.pdt_dep_nodba_cross_schema_privileges(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_dep_nodba_external_references_connection_id ON oracle_migration.pdt_dep_nodba_external_references(connection_id)")
            
            # Indici ora2pg ptd_ (con target_schema)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_ptd_ora2pg_estimates_connection_id ON oracle_migration.ptd_ora2pg_estimates(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_ptd_ora2pg_estimates_target_schema ON oracle_migration.ptd_ora2pg_estimates(target_schema)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_ptd_ora2pg_object_summary_connection_id ON oracle_migration.ptd_ora2pg_object_summary(connection_id)")
            
            # Indici dimensioni DBA pdt_sizes_dba_
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_dba_database_size_connection_id ON oracle_migration.pdt_sizes_dba_database_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_dba_tablespace_size_connection_id ON oracle_migration.pdt_sizes_dba_tablespace_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_dba_schema_size_connection_id ON oracle_migration.pdt_sizes_dba_schema_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_dba_table_size_connection_id ON oracle_migration.pdt_sizes_dba_table_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_dba_index_size_connection_id ON oracle_migration.pdt_sizes_dba_index_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_dba_segment_size_connection_id ON oracle_migration.pdt_sizes_dba_segment_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_dba_code_lines_connection_id ON oracle_migration.pdt_sizes_dba_code_lines(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_dba_code_stats_connection_id ON oracle_migration.pdt_sizes_dba_code_stats(connection_id)")
            
            # Indici dimensioni NON-DBA pdt_sizes_nodba_
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_nodba_database_size_connection_id ON oracle_migration.pdt_sizes_nodba_database_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_nodba_tablespace_size_connection_id ON oracle_migration.pdt_sizes_nodba_tablespace_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_nodba_schema_size_connection_id ON oracle_migration.pdt_sizes_nodba_schema_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_nodba_table_size_connection_id ON oracle_migration.pdt_sizes_nodba_table_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_nodba_index_size_connection_id ON oracle_migration.pdt_sizes_nodba_index_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_nodba_segment_size_connection_id ON oracle_migration.pdt_sizes_nodba_segment_size(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_nodba_code_lines_connection_id ON oracle_migration.pdt_sizes_nodba_code_lines(connection_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_pdt_sizes_nodba_code_stats_connection_id ON oracle_migration.pdt_sizes_nodba_code_stats(connection_id)")
            
            conn.commit()
            cursor.close()
            conn.close()
            print("  ✅ Schema database con prefissi completi e query tablespace aggiornate creato/aggiornato con successo")
            
        except Exception as e:
            print(f"  ❌ Errore creazione schema database: {e}")
            raise
    
    def get_or_create_connection_id(self, db_config, is_dba):
        """🆕 Ottiene o crea l'ID della connessione nella tabella pdt_connections con schema"""
        try:
            conn = psycopg2.connect(**self.pg_config)
            cursor = conn.cursor()
            
            # Verifica se la connessione esiste già
            cursor.execute("""
                SELECT id FROM oracle_migration.pdt_connections 
                WHERE connection_name = %s
            """, (db_config['connection_name'],))
            
            result = cursor.fetchone()
            
            analyze_all = db_config.get('analyze_all_schemas', True)
            schema = db_config.get('schema', None)
            
            if result:
                connection_id = result[0]
                # 🔧 Aggiorna i dati della connessione includendo schema
                cursor.execute("""
                    UPDATE oracle_migration.pdt_connections 
                    SET dsn = %s, username = %s, schema = %s, description = %s, is_dba = %s, analyze_all_schemas = %s, updated_date = CURRENT_TIMESTAMP
                    WHERE id = %s
                """, (
                    db_config.get('dsn', ''),
                    db_config['user'],
                    schema,
                    db_config.get('description', ''),
                    is_dba,
                    analyze_all,
                    connection_id
                ))
                schema_info = f", Schema: {schema}" if schema else ""
                print(f"    > Connessione aggiornata: ID {connection_id} (DBA: {'SÌ' if is_dba else 'NO'}{schema_info})")
            else:
                # 🔧 Crea nuova connessione con schema
                cursor.execute("""
                    INSERT INTO oracle_migration.pdt_connections 
                    (connection_name, dsn, username, schema, description, is_dba, analyze_all_schemas)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    db_config['connection_name'],
                    db_config.get('dsn', ''),
                    db_config['user'],
                    schema,
                    db_config.get('description', ''),
                    is_dba,
                    analyze_all
                ))
                connection_id = cursor.fetchone()[0]
                schema_info = f", Schema: {schema}" if schema else ""
                print(f"    > Nuova connessione creata: ID {connection_id} (DBA: {'SÌ' if is_dba else 'NO'}{schema_info})")
            
            conn.commit()
            cursor.close()
            conn.close()
            return connection_id
            
        except Exception as e:
            print(f"    ❌ Errore gestione connessione: {e}")
            raise
    
    def cleanup_existing_data(self, connection_id, is_dba):
        """🆕 Cancella i dati esistenti per una connessione dalle tabelle con prefissi DBA/NON-DBA"""
        try:
            conn = psycopg2.connect(**self.pg_config)
            cursor = conn.cursor()
            
            # 🆕 Lista delle tabelle da pulire con prefissi DBA/NON-DBA
            tables_to_clean = [
                # Tabelle ora2pg (ptd_) - sempre
                'ptd_ora2pg_object_summary',
                'ptd_ora2pg_estimates'
            ]
            
            # Aggiungi tabelle dipendenze in base al tipo utente
            if is_dba:
                tables_to_clean.extend([
                    'pdt_dep_dba_external_references',
                    'pdt_dep_dba_cross_schema_privileges',
                    'pdt_dep_dba_db_links',
                    'pdt_dep_dba_dependencies',
                    'pdt_sizes_dba_code_lines',
                    'pdt_sizes_dba_code_stats',
                    'pdt_sizes_dba_segment_size',
                    'pdt_sizes_dba_index_size',
                    'pdt_sizes_dba_table_size',
                    'pdt_sizes_dba_schema_size',
                    'pdt_sizes_dba_tablespace_size',
                    'pdt_sizes_dba_database_size'
                ])
            else:
                tables_to_clean.extend([
                    'pdt_dep_nodba_external_references',
                    'pdt_dep_nodba_cross_schema_privileges',
                    'pdt_dep_nodba_db_links',
                    'pdt_dep_nodba_dependencies',
                    'pdt_sizes_nodba_code_lines',
                    'pdt_sizes_nodba_code_stats',
                    'pdt_sizes_nodba_segment_size',
                    'pdt_sizes_nodba_index_size',
                    'pdt_sizes_nodba_table_size',
                    'pdt_sizes_nodba_schema_size',
                    'pdt_sizes_nodba_tablespace_size',
                    'pdt_sizes_nodba_database_size'
                ])
            
            total_deleted = 0
            for table in tables_to_clean:
                cursor.execute(f"""
                    DELETE FROM oracle_migration.{table} 
                    WHERE connection_id = %s
                """, (connection_id,))
                deleted_count = cursor.rowcount
                total_deleted += deleted_count
                if deleted_count > 0:
                    print(f"    > Eliminati {deleted_count} record da {table}")
            
            conn.commit()
            cursor.close()
            conn.close()
            
            if total_deleted > 0:
                print(f"    > Totale record eliminati: {total_deleted}")
            else:
                print(f"    > Nessun dato esistente da eliminare")
                
        except Exception as e:
            print(f"    ❌ Errore pulizia dati esistenti: {e}")
            raise
    
    def save_to_postgresql(self, all_results):
        """🆕 Salva tutti i risultati nel database PostgreSQL con schema_name = schema configurato"""
        try:
            print(f"\n💾 Creazione/aggiornamento schema database con query tablespace aggiornate...")
            self.create_database_schema()
            
            total_records = 0
            
            # Elabora ogni database
            for db_key, results in all_results.items():
                if results.get('error'):
                    print(f"    > Saltato {db_key} per errore: {results['error']}")
                    continue
                
                connection_name = results.get('connection_name')
                if not connection_name:
                    print(f"    ⚠️  Saltato {db_key}: connection_name mancante")
                    continue
                
                print(f"\n    📊 Elaborazione {connection_name}...")
                
                # Ottieni connection_id e privilegi DBA
                db_config = next((conn for conn in self.oracle_connections 
                                if conn['connection_name'] == connection_name), None)
                if not db_config:
                    print(f"    ⚠️  Configurazione non trovata per {connection_name}")
                    continue
                
                # Determina se è DBA dai dati raccolti
                is_dba = results.get('is_dba', False)
                
                connection_id = self.get_or_create_connection_id(db_config, is_dba)
                
                # Pulisci dati esistenti
                self.cleanup_existing_data(connection_id, is_dba)
                
                # Connessione per inserimenti
                conn = psycopg2.connect(**self.pg_config)
                cursor = conn.cursor()
                
                # 🔧 schema_name = schema configurato nel JSON, non l'utente
                target_schema = db_config.get('schema', results.get('schema', 'UNKNOWN'))
                
                # 🆕 INSERISCI DIPENDENZE CON PREFISSI DBA/NON-DBA
                dep_prefix = "pdt_dep_dba_" if is_dba else "pdt_dep_nodba_"
                print(f"      > Inserimento dipendenze con prefisso: {dep_prefix}")
                
                for dep in results.get('dependencies', []):
                    cursor.execute(f"""
                        INSERT INTO oracle_migration.{dep_prefix}dependencies 
                        (connection_id, source_owner, source_name, source_type, 
                         target_owner, target_name, target_type, db_link)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (connection_id, *dep))
                    total_records += 1
                
                for link in results.get('db_links', []):
                    cursor.execute(f"""
                        INSERT INTO oracle_migration.{dep_prefix}db_links 
                        (connection_id, owner, db_link, username, host)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (connection_id, *link))
                    total_records += 1
                
                for priv in results.get('cross_schema_privs', []):
                    cursor.execute(f"""
                        INSERT INTO oracle_migration.{dep_prefix}cross_schema_privileges 
                        (connection_id, grantor, grantee, table_schema, table_name, privilege)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (connection_id, *priv))
                    total_records += 1
                
                for ref in results.get('external_references', []):
                    cursor.execute(f"""
                        INSERT INTO oracle_migration.{dep_prefix}external_references 
                        (connection_id, synonym_owner, synonym_name, referenced_owner, referenced_object, db_link)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (connection_id, *ref))
                    total_records += 1
                
                # 🆕 INSERISCI ORA2PG CON schema_name = schema configurato
                if 'ora2pg_metrics' in results and 'ora2pg_object_summary' in results['ora2pg_metrics']:
                    for obj in results['ora2pg_metrics']['ora2pg_object_summary']:
                        cursor.execute("""
                            INSERT INTO oracle_migration.ptd_ora2pg_object_summary (
                                connection_id, schema_name, object_name, object_number, 
                                invalid_count, estimated_cost, comments, details,
                                detail_type, procedure_name, procedure_cost
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            connection_id,
                            target_schema,  # 🔧 Usa schema configurato
                            obj['object_name'],
                            obj['object_number'],
                            obj['invalid_count'],
                            obj['estimated_cost'],
                            obj['comments'],
                            obj['details'],
                            obj.get('detail_type', 'MAIN'),
                            obj.get('procedure_name'),
                            obj.get('procedure_cost')
                        ))
                        total_records += 1
                    print(f"      > Inseriti {len(results['ora2pg_metrics']['ora2pg_object_summary'])} record ptd_ora2pg_object_summary (schema: {target_schema})")
                
                if 'ora2pg_metrics' in results:
                    # 🆕 SALVA INFO COMPLETE ORA2PG CON schema_name = schema configurato
                    analyzed_schemas_str = ', '.join(results['ora2pg_metrics'].get('analyzed_schemas', [target_schema]))
                    configured_target_schema = results['ora2pg_metrics'].get('target_schema', 'auto')
                    
                    cursor.execute("""
                        INSERT INTO oracle_migration.ptd_ora2pg_estimates 
                        (connection_id, schema_name, total_cost, migration_level, 
                         analyzed_schemas, target_schema, dba_mode, reports_count, metrics)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        connection_id, 
                        target_schema,  # 🔧 Usa schema configurato
                        results['ora2pg_metrics'].get('total_cost', 0),
                        results['ora2pg_metrics'].get('migration_level', 'Unknown'),
                        analyzed_schemas_str,
                        configured_target_schema,
                        results['ora2pg_metrics'].get('dba_mode', False),
                        results['ora2pg_metrics'].get('reports_count', 1),
                        json.dumps(results['ora2pg_metrics'])
                    ))
                    total_records += 1
                    
                    print(f"      > Inserito record ptd_ora2pg_estimates (schema: {target_schema}, target: {configured_target_schema})")
                
                # 🆕 INSERISCI DATI DIMENSIONI CON PREFISSI DBA/NON-DBA E QUERY TABLESPACE AGGIORNATE
                if self.analyze_sizes and 'size_data' in results:
                    size_data = results['size_data']
                    
                    # Determina prefisso in base a privilegi DBA
                    size_prefix = "pdt_sizes_dba_" if is_dba else "pdt_sizes_nodba_"
                    
                    print(f"      > Inserimento dimensioni con prefisso: {size_prefix}")
                    
                    # Database Size
                    for db_size in size_data.get('database_size', []):
                        cursor.execute(f"""
                            INSERT INTO oracle_migration.{size_prefix}database_size 
                            (connection_id, metric_type, object_name, size_gb, size_mb, size_bytes, file_count)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (connection_id, *db_size))
                        total_records += 1
                    
                    # 🔧 Tablespace Size con colonne aggiornate
                    for ts_size in size_data.get('tablespace_size', []):
                        if is_dba:
                            # DBA: query dettagliata con tutte le colonne
                            cursor.execute(f"""
                                INSERT INTO oracle_migration.{size_prefix}tablespace_size 
                                (connection_id, tablespace_name, status, type, allocated_gb, allocated_mb, allocated_bytes,
                                 used_gb, used_mb, used_bytes, free_gb, free_mb, free_bytes,
                                 pct_used, pct_free, datafile_count, segment_count)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (connection_id, *ts_size))
                        else:
                            # NON-DBA: query aggregata
                            cursor.execute(f"""
                                INSERT INTO oracle_migration.{size_prefix}tablespace_size 
                                (connection_id, tablespace_name, used_gb, used_mb, used_bytes, file_count, status)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """, (connection_id, *ts_size))
                        total_records += 1
                    
                    # Schema Size
                    for schema_size in size_data.get('schema_size', []):
                        cursor.execute(f"""
                            INSERT INTO oracle_migration.{size_prefix}schema_size 
                            (connection_id, owner, size_gb, size_mb, size_bytes, segment_count)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (connection_id, *schema_size))
                        total_records += 1
                    
                    # Table Size
                    for table_size in size_data.get('table_size', []):
                        cursor.execute(f"""
                            INSERT INTO oracle_migration.{size_prefix}table_size 
                            (connection_id, owner, table_name, segment_type, tablespace_name, size_gb, size_mb, size_bytes, blocks, extents)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (connection_id, *table_size))
                        total_records += 1
                    
                    # Index Size
                    for index_size in size_data.get('index_size', []):
                        cursor.execute(f"""
                            INSERT INTO oracle_migration.{size_prefix}index_size 
                            (connection_id, owner, index_name, segment_type, tablespace_name, size_gb, size_mb, size_bytes, blocks, extents)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (connection_id, *index_size))
                        total_records += 1
                    
                    # Segment Size
                    for segment_size in size_data.get('segment_size', []):
                        cursor.execute(f"""
                            INSERT INTO oracle_migration.{size_prefix}segment_size 
                            (connection_id, owner, segment_name, segment_type, tablespace_name, size_gb, size_mb, size_bytes, blocks, extents, initial_extent, next_extent, max_extents)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (connection_id, *segment_size))
                        total_records += 1
                    
                    # Code Lines (limitato per performance)
                    code_lines = size_data.get('code_lines', [])
                    if len(code_lines) <= 10000:
                        for code_line in code_lines:
                            cursor.execute(f"""
                                INSERT INTO oracle_migration.{size_prefix}code_lines 
                                (connection_id, owner, object_name, object_type, char_length, byte_length, line_number, line_text)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            """, (connection_id, *code_line))
                            total_records += 1
                    else:
                        print(f"      ⚠️  Troppi record code_lines ({len(code_lines)}), inserimento saltato per performance")
                    
                    # Code Stats
                    for code_stats in size_data.get('code_stats', []):
                        cursor.execute(f"""
                            INSERT INTO oracle_migration.{size_prefix}code_stats 
                            (connection_id, owner, object_name, object_type, total_lines, total_chars, total_bytes, first_line, last_line)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (connection_id, *code_stats))
                        total_records += 1
                    
                    print(f"      > Inseriti dati dimensioni per {connection_name} (modalità: {'DBA' if is_dba else 'NON-DBA'})")
                
                conn.commit()
                cursor.close()
                conn.close()
                
                print(f"      ✅ Dati salvati per {connection_name}")
            
            print(f"\n> Dati salvati in PostgreSQL con schema completo aggiornato! ({total_records} record totali)")
            print(f"  📋 Prefissi utilizzati: pdt_dep_dba_/pdt_dep_nodba_ (dipendenze), pdt_sizes_dba_/pdt_sizes_nodba_ (dimensioni), ptd_ (ora2pg)")
            print(f"  🔧 Query tablespace: DBA (dettagli completi), NON-DBA (aggregato per tablespace)")
            print(f"  📊 schema_name nelle tabelle ora2pg = schema configurato nel JSON")
            print(f"  🔧 MODIFICHE: File Excel specifici disattivati, naming con schema invece di username")
            
        except Exception as e:
            print(f"\n❌ Errore nel salvataggio su PostgreSQL: {e}")
            print("I dati sono comunque stati salvati nei file di output")
            import traceback
            traceback.print_exc()
    
    def analyze_database(self, db_config):
        """🆕 Analizza un singolo database Oracle con rilevamento DBA e query separate - MODIFICATO NAMING"""
        connection_name = db_config['connection_name']
        
        # Verifica se DSN è presente
        if 'dsn' not in db_config or not db_config['dsn']:
            print(f"\n{'='*70}")
            print(f"⚠️  SALTATO - Database: {connection_name}")
            print(f"🔗 Motivo: DSN mancante nella configurazione")
            print(f"{'='*70}")
            return {
                'schema': db_config['user'],
                'connection_name': connection_name,
                'dsn': 'N/A',
                'error': 'DSN mancante nella configurazione'
            }
        
        db_name = f"{connection_name}_{db_config['user']}@{db_config['dsn']}"
        
        print(f"\n{'='*70}")
        print(f"📊 Analisi database: {connection_name}")
        print(f"🔗 Connessione: {db_config['user']}@{db_config['dsn']}")
        if 'description' in db_config:
            print(f"📝 Descrizione: {db_config['description']}")
        if 'schema' in db_config:
            print(f"🎯 Schema ora2pg configurato: {db_config['schema']}")
        print(f"{'='*70}")
        
        results = {
            'schema': db_config['user'],
            'connection_name': connection_name,
            'dsn': db_config['dsn']
        }
        
        try:
            # Connessione Oracle
            connection = self.get_db_connection(
                db_config['dsn'],
                db_config['user'],
                db_config['password']
            )
            
            print("  ✅ Connessione Oracle riuscita")
            
            # 🆕 VERIFICA PRIVILEGI DBA
            is_dba = self.check_dba_privileges(connection, db_config)
            results['is_dba'] = is_dba
            
            # 🆕 ESTRAI DIPENDENZE CON QUERY SPECIFICHE PER DBA/NON-DBA
            print("  📊 Estrazione dipendenze database...")
            oracle_data = self.get_oracle_dependencies(connection, is_dba, db_config)
            results.update(oracle_data)
            print(f"  ✅ Dipendenze estratte con successo (modalità: {'DBA' if is_dba else 'NON-DBA'})")
            
            # 🆕 ESTRAI DIMENSIONI (se abilitato) CON QUERY SPECIFICHE PER DBA/NON-DBA
            if self.analyze_sizes:
                print("  📏 Estrazione dimensioni database con query tablespace aggiornate...")
                size_data = self.get_oracle_sizes(connection, is_dba, db_config)
                results['size_data'] = size_data
                oracle_data['size_data'] = size_data  # Aggiungi anche a oracle_data per Excel
                print(f"  ✅ Dimensioni estratte con successo (modalità: {'DBA (dettagli completi)' if is_dba else 'NON-DBA (aggregato)'})")
            
            # ==========================================
            # SEZIONE GENERAZIONE FILE DI OUTPUT - MODIFICATA
            # ==========================================
            print("  📄 Generazione file di output...")
            
            # 🔧 DETERMINA SCHEMA NAME PER NAMING
            schema_name = db_config.get('schema', db_config['user'])  # Usa schema se configurato, altrimenti user
            
            # === GENERAZIONE FILE CSV (DISABILITATA) ===
            if self.generate_csv:
                dba_suffix = "_dba" if is_dba else "_nodba"
                print(f"    📋 Creazione file CSV con suffisso: {dba_suffix}")
                
                # ❌ DISATTIVATO - Dependencies CSV
                # if oracle_data['dependencies']:
                #     self.save_to_csv(...)
                
                if oracle_data['db_links']:
                    self.save_to_csv(
                        oracle_data['db_links'],
                        f"{connection_name}_dblinks{dba_suffix}_{schema_name}.csv",  # 🔧 SCHEMA INVECE DI USER
                        ['OWNER', 'DB_LINK', 'USERNAME', 'HOST']
                    )
                
                # ❌ DISATTIVATO - Objects CSV  
                # if oracle_data['object_summary']:
                #     self.save_to_csv(...)
                
                # ❌ DISATTIVATO - Cross Schema Privs CSV
                # if oracle_data['cross_schema_privs']:
                #     self.save_to_csv(...)
                
                # ❌ DISATTIVATO - External References CSV
                # if oracle_data['external_references']:
                #     self.save_to_csv(...)
            
            # === GENERAZIONE FILE EXCEL (ABILITATA CON MODIFICHE) ===
            if self.generate_excel:
                dba_suffix = "_dba" if is_dba else "_nodba"
                print(f"    📊 Creazione file Excel con suffisso: {dba_suffix}")
                
                # ❌ DISATTIVATO - Dependencies Excel
                # if oracle_data['dependencies']:
                #     self.save_to_excel(...)
                
                if oracle_data['db_links']:
                    self.save_to_excel(
                        oracle_data['db_links'],
                        f"{connection_name}_dblinks{dba_suffix}_{schema_name}.xlsx",  # 🔧 SCHEMA INVECE DI USER
                        ['OWNER', 'DB_LINK', 'USERNAME', 'HOST'],
                        "DB_Links"
                    )
                
                # ❌ DISATTIVATO - Objects Excel
                # if oracle_data['object_summary']:
                #     self.save_to_excel(...)
                
                # ❌ DISATTIVATO - Cross Schema Privs Excel
                # if oracle_data['cross_schema_privs']:
                #     self.save_to_excel(...)
                
                # ❌ DISATTIVATO - External References Excel
                # if oracle_data['external_references']:
                #     self.save_to_excel(...)
                
                # === REPORT EXCEL COMPLETO ===
                print("    📈 Creazione report Excel completo...")
                self.save_combined_excel_report(oracle_data, connection_name, schema_name, is_dba)  # 🔧 SCHEMA INVECE DI USER
                
                # 🆕 === REPORT EXCEL DIMENSIONI con headers tablespace aggiornati ===
                if self.analyze_sizes and 'size_data' in results:
                    print("    📏 Creazione report Excel dimensioni con query tablespace aggiornate...")
                    self.save_sizes_excel_report(results['size_data'], connection_name, schema_name)  # 🔧 SCHEMA INVECE DI USER
            
            print("  ✅ File di output generati con successo")
            # ==========================================
            # FINE SEZIONE GENERAZIONE FILE DI OUTPUT
            # ==========================================
            
            connection.close()
            
            # 🆕 ESEGUI ORA2PG CON GESTIONE SCHEMA CONFIGURABILE
            print(f"  📊 Esecuzione analisi ora2pg (modalità: {self.ora2pg_output_mode})...")
            ora2pg_results = self.run_ora2pg_analysis(
                db_config['dsn'],
                db_config['user'],
                db_config['password'],
                connection_name,
                is_dba,
                db_config
            )
            if ora2pg_results:
                results['ora2pg_metrics'] = ora2pg_results
                target_schema = ora2pg_results.get('target_schema', 'auto')
                mode_desc = f"schema: {target_schema}" if target_schema != 'auto' else ("DB completo" if ora2pg_results.get('dba_mode') else "solo schema")
                print(f"  ✅ Analisi ora2pg completata ({mode_desc}) - Costo: {ora2pg_results.get('total_cost', 'N/A')}")
            
            print(f"  🎉 Analisi {connection_name} completata con successo")
            
        except Exception as e:
            error_msg = f"Errore durante l'analisi di {connection_name}: {str(e)}"
            print(f"  ❌ {error_msg}")
            results['error'] = error_msg
            import traceback
            traceback.print_exc()
            
        return results

    def run_analysis(self):
        """Esegue l'analisi per tutti i database configurati"""
        print(f"\n🚀 INIZIO ANALISI MULTI-DATABASE")
        print(f"📅 Data/ora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"📁 Output directory: {self.output_dir}")
        print(f"🔢 Database da analizzare: {len(self.oracle_connections)}")
        print(f"📊 Formato output: {'Excel ✅' if self.generate_excel else ''}{'CSV ✅' if self.generate_csv else ''}")
        print(f"📋 Output ora2pg: {self.ora2pg_output_mode}")
        print(f"📏 Analisi dimensioni: {'Abilitata ✅' if self.analyze_sizes else 'Disabilitata ❌'}")
        print(f"🗄️  Database PostgreSQL: Prefissi completi (pdt_dep_dba_/pdt_dep_nodba_, pdt_sizes_dba_/pdt_sizes_nodba_, ptd_)")
        print(f"🔍 Rilevamento privilegi DBA: Automatico/Configurazione")
        print(f"🎯 Query differenziate: DBA vs NON-DBA per dipendenze e dimensioni")
        print(f"🔧 ORA2PG: Schema configurabile per connessione, fallback logica DBA/NON-DBA")
        print(f"📊 Query tablespace: DBA (dettagli completi), NON-DBA (aggregato per tablespace)")
        print(f"🗃️  schema_name nelle tabelle ora2pg = schema configurato nel JSON")
        print(f"🔧 MODIFICHE: File Excel specifici disattivati, naming con schema invece di username")
        
        all_results = {}
        successful_analyses = 0
        failed_analyses = 0
        
        # Analizza ogni database
        for i, db_config in enumerate(self.oracle_connections, 1):
            print(f"\n📋 Elaborazione {i}/{len(self.oracle_connections)}")
            
            try:
                results = self.analyze_database(db_config)
                connection_name = db_config['connection_name']
                db_key = f"{connection_name}_{db_config['user']}@{db_config.get('dsn', 'N/A')}"
                all_results[db_key] = results
                
                if results.get('error'):
                    failed_analyses += 1
                else:
                    successful_analyses += 1
                    
            except Exception as e:
                print(f"  ❌ Errore critico per {db_config['connection_name']}: {str(e)}")
                failed_analyses += 1
        
        # Salva report riassuntivo
        print(f"\n📄 Generazione report riassuntivo...")
        self.save_summary_report(all_results)
        
        # Salva tutto in PostgreSQL
        print(f"\n💾 Salvataggio dati in PostgreSQL con query tablespace aggiornate...")
        self.save_to_postgresql(all_results)
        
        # Report finale
        print(f"\n{'='*70}")
        print(f"🏁 ANALISI COMPLETATA!")
        print(f"{'='*70}")
        print(f"✅ Database analizzati con successo: {successful_analyses}")
        print(f"❌ Database con errori: {failed_analyses}")
        print(f"📁 Risultati salvati in: {self.output_dir}")
        print(f"📋 File di configurazione: {self.config_file}")
        print(f"📊 Formato output: {'Excel ✅' if self.generate_excel else ''}{'CSV ✅' if self.generate_csv else ''}")
        print(f"📋 Output ora2pg: {self.ora2pg_output_mode}")
        print(f"📏 Analisi dimensioni: {'Abilitata ✅' if self.analyze_sizes else 'Disabilitata ❌'}")
        print(f"🗄️  Database PostgreSQL: Prefissi completi (pdt_dep_dba_/pdt_dep_nodba_, pdt_sizes_dba_/pdt_sizes_nodba_, ptd_)")
        print(f"🔍 Rilevamento privilegi DBA: Implementato con query differenziate")
        print(f"🎯 Query ottimizzate: DBA (tutti gli schemi) vs NON-DBA (solo schema utente)")
        print(f"🔧 ORA2PG: Schema configurabile per connessione, fallback logica DBA/NON-DBA")
        print(f"📊 Query tablespace: DBA (dettagli completi), NON-DBA (aggregato per tablespace)")
        print(f"🗃️  schema_name nelle tabelle ora2pg = schema configurato nel JSON")
        print(f"🔧 Correzione: Usato 'owner' invece di 'table_schema' nelle query dba_tab_privs/all_tab_privs")
        print(f"🔧 MODIFICHE APPLICATE: File Excel specifici disattivati, naming con schema invece di username")
        print(f"{'='*70}")
        
        # Lista file generati
        print("\n📂 File generati:")
        for file in sorted(os.listdir(self.output_dir)):
            print(f"  - {file}")


def main():
    """Funzione principale"""
    import argparse
    
    print("🎯 Oracle Multi-Database Dependency Analyzer")
    print("📋 Versione con query tablespace aggiornate, schema configurabile ora2pg, correzione query dba_tab_privs")
    print("🔧 MODIFICHE: File Excel specifici disattivati, naming con schema invece di username")
    
    parser = argparse.ArgumentParser(description='Oracle Multi-Database Dependency Analyzer')
    parser.add_argument(
        '--config', 
        default='oracle_connections.json',
        help='File di configurazione JSON (default: oracle_connections.json)'
    )
    parser.add_argument(
        '--csv', 
        action='store_true',
        help='Abilita generazione file CSV (default: disabilitato)'
    )
    parser.add_argument(
        '--no-excel', 
        action='store_true',
        help='Disabilita generazione file Excel (default: abilitato)'
    )
    parser.add_argument(
        '--ora2pg-mode',
        choices=['html_only', 'html_and_txt'],
        help='Modalità output ora2pg: html_only o html_and_txt (default: da config)'
    )
    parser.add_argument(
        '--no-sizes',
        action='store_true',
        help='Disabilita analisi dimensioni (default: abilitata)'
    )
    
    args = parser.parse_args()
    
    try:
        # Inizializza analyzer
        print(f"🔧 Inizializzazione analyzer...")
        analyzer = OracleMultiDatabaseAnalyzer(args.config)
        
        # Override configurazioni da parametri command line
        if args.csv:
            analyzer.generate_csv = True
            print("📋 CSV abilitato da parametro command line")
            
        if args.no_excel:
            analyzer.generate_excel = False
            print("📊 Excel disabilitato da parametro command line")
            
        if args.ora2pg_mode:
            analyzer.ora2pg_output_mode = args.ora2pg_mode
            print(f"📋 Modalità ora2pg impostata da command line: {args.ora2pg_mode}")
            
        if args.no_sizes:
            analyzer.analyze_sizes = False
            print("📏 Analisi dimensioni disabilitata da parametro command line")
        
        # Esegui analisi
        analyzer.run_analysis()
        
    except KeyboardInterrupt:
        print("\n\n⏹️  Analisi interrotta dall'utente")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Errore critico: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()